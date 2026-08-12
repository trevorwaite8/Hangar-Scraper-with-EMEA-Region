"""
Safespill Hangar News Scraper  (v2 — 2026-08-12)
================================================================================
Runs weekly. Searches for hangar / aviation-facility / fire-protection project
news and public procurement notices worldwide, then emails a CUMULATIVE Excel
report to the Safespill team.

WHAT CHANGED IN v2 (vs. the 8/12/26 "prior to updates" version)
--------------------------------------------------------------------------------
1. NEWS, NOT THE WHOLE WEB.  SerpAPI is now called with tbm=nws and
   tbs=qdr:w, so results are actual news articles from the past week.
   v1 used engine=google with no date filter, which is why the 8/3 report
   was full of vendor brochures, YouTube, Reddit, Facebook and Wikipedia.

2. LOCATION ACCURACY.  detect_location() was rewritten (see the GEOGRAPHY
   section).  Title is checked first, script/style/nav junk is stripped before
   any text analysis, a 400-entry airport / air-base gazetteer was added, the
   publisher's domain and TLD are used as evidence, and ambiguous names
   ("Georgia", "Washington") are disambiguated instead of guessed.
   Rows that still cannot be placed go to a "Needs Review" tab rather than
   being silently dumped into the USA tab.

3. FIVE TABS instead of two: USA, Canada, ANZ, EMEA, Needs Review.

4. CUMULATIVE HISTORY.  history.csv is committed back to the repo each run.
   New rows are prepended above prior weeks, deduped across weeks, and rows
   older than HISTORY_RETENTION_DAYS are dropped.

5. RELEVANCE SCORING + DOMAIN BLOCKLIST to hit the 30-50 new rows/week target.

6. WORKING PROCUREMENT FEEDS.  AusTender, CanadaBuys and NZ GETS all work now
   and none of them require an API key (see the PROCUREMENT section).

7. EMAIL VIA MICROSOFT GRAPH so the report sends from a safespill.com address.
   Falls back to SMTP automatically if Graph secrets are not configured yet.

--------------------------------------------------------------------------------
CREDENTIALS
--------------------------------------------------------------------------------
REQUIRED
  SERPAPI_KEY       Free tier = 250 searches/month, 50/hour.
                    This script uses SEARCH_BUDGET_PER_RUN (default 40) per
                    run, i.e. ~174/month at one run per week. Check quota at
                    https://serpapi.com/dashboard
  SAM_API_KEY       Rotates every 90 days. SAM.gov emails a reminder ~10 days
                    out. Retrieve from sam.gov/workspace/profile/account-details
                    (Public API Key -> eye icon -> one-time password -> copy),
                    then update the GitHub secret.

EMAIL — pick ONE of these two blocks
  (A) Microsoft Graph  [preferred: no password to rotate, survives Microsoft's
      retirement of Basic Auth for SMTP AUTH]
        GRAPH_TENANT_ID       Entra directory (tenant) ID
        GRAPH_CLIENT_ID       App registration's application (client) ID
        GRAPH_CLIENT_SECRET   Client secret value
        GRAPH_SENDER          Mailbox to send as, e.g.
                              hangar-report@safespill.com
  (B) SMTP  [legacy fallback]
        SMTP_HOST / SMTP_PORT / SMTP_USER / SMTP_PASSWORD

  REPORT_RECIPIENT   Comma-separated recipient list.

NO KEY NEEDED
  AusTender OCDS, CanadaBuys open data, NZ GETS RSS, TED Europa.
  If you previously tried to "register for an API" on those sites and got
  nowhere, that is because there is nothing to register for.

If the weekly log shows every call from one source failing (e.g. all SAM.gov
calls return 401/403), a rotated or stale credential is the likely cause.
"""

from __future__ import annotations

import base64
import csv
import datetime
import io
import json
import logging
import os
import re
import time
from urllib.parse import urlparse, urlunparse

import openpyxl
import requests
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --- Logging ------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    handlers=[logging.StreamHandler()],
)
log = logging.getLogger(__name__)


# ==============================================================================
# CONFIGURATION
# ==============================================================================
#
# Read lazily via env() rather than at import time. v1 did
# SERPAPI_KEY = os.environ["SERPAPI_KEY"] at module level, which made the file
# impossible to import for testing without the full production environment.

def env(name: str, default: str = "") -> str:
    return os.environ.get(name, default)


# How far back news results may be dated. Rows outside this window are dropped.
LOOKBACK_DAYS = 10          # 10 not 7: gives slack for slow-indexing sites
SAM_LOOKBACK_DAYS = 30      # contracts post infrequently
PROCUREMENT_LOOKBACK_DAYS = 14

# Cumulative history
HISTORY_FILE = env("HISTORY_FILE", "history.csv")
HISTORY_RETENTION_DAYS = 365        # rolling 12 months

# SerpAPI budget. Free tier is 250 searches/month; at ~4.33 weekly runs per
# month, 40/run = ~174/month and leaves room for manual test runs.
SEARCH_BUDGET_PER_RUN = int(env("SEARCH_BUDGET_PER_RUN", "40"))

# Target volume of NEW rows per week. Used to auto-tighten the relevance
# threshold when a week over-delivers.
TARGET_NEW_ROWS_MIN = 30
TARGET_NEW_ROWS_MAX = 50

HEADERS_UA = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/126.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}


# ==============================================================================
# SEARCH QUERIES
# ==============================================================================
#
# Budget-aware design. CORE_QUERIES run every week. ROTATING_QUERIES is a
# larger pool from which a different slice runs each week (keyed off the ISO
# week number), so coverage broadens over time at no extra API cost.
#
# Regional terms are combined with OR inside a single query — one search
# instead of four. That is what buys back enough budget to cover ANZ and NZ.

CORE_QUERIES = [
    # New construction / development
    "aircraft hangar construction",
    "new hangar groundbreaking",
    "hangar expansion project",
    "hangar construction contract awarded",
    "airport hangar development",
    # FBO / corporate aviation
    "FBO hangar development",
    "business jet hangar construction",
    # MRO
    "MRO facility construction",
    "aircraft maintenance hangar project",
    "widebody maintenance hangar",
    # Military
    "military hangar construction contract",
    "Air Force hangar construction",
    "Navy aircraft hangar construction",
    # Fire protection — the direct Safespill signal
    "hangar fire suppression system",
    "aircraft hangar fire protection",
    "NFPA 409 hangar foam",
    "AFFF replacement hangar",
    "fluorine free foam hangar transition",
    # Retrofit
    "hangar renovation contract",
    "hangar retrofit fire suppression",
    # Consolidated regional sweeps (OR keeps these to one search each)
    "hangar construction Dubai OR UAE OR \"Saudi Arabia\" OR Qatar",
    "hangar construction UK OR Germany OR France OR Poland",
    "hangar construction Australia OR \"New Zealand\"",
    "hangar OR MRO construction Nigeria OR Kenya OR \"South Africa\" OR Egypt",
]

ROTATING_QUERIES = [
    "paint hangar construction",
    "engine test facility construction airport",
    "helicopter hangar construction",
    "cargo facility hangar construction",
    "aircraft completions center construction",
    "airport capital improvement hangar",
    "hangar lease development airport authority",
    "MILCON hangar award",
    "National Guard hangar construction",
    "aircraft hangar design build contract",
    "hangar foam system upgrade",
    "airport fire suppression upgrade",
    "hangar construction India OR Singapore OR Japan",
    "hangar construction Canada OR Ontario OR Alberta",
    "defence hangar infrastructure NATO",
    "aircraft shelter construction air base",
    "hangar construction Turkey OR Greece OR Italy",
    "aviation facility investment announcement",
]


def queries_for_this_week(week_iso: int) -> list[str]:
    """
    Build this run's query list: all CORE_QUERIES plus a rotating slice of
    ROTATING_QUERIES, capped at SEARCH_BUDGET_PER_RUN.
    """
    slots = max(0, SEARCH_BUDGET_PER_RUN - len(CORE_QUERIES))
    if not ROTATING_QUERIES or slots == 0:
        return list(CORE_QUERIES)[:SEARCH_BUDGET_PER_RUN]
    start = (week_iso * slots) % len(ROTATING_QUERIES)
    rotating = [ROTATING_QUERIES[(start + i) % len(ROTATING_QUERIES)]
                for i in range(min(slots, len(ROTATING_QUERIES)))]
    return list(CORE_QUERIES) + rotating


# ==============================================================================
# RELEVANCE FILTERING
# ==============================================================================
#
# v1 had no filtering at all, which is how facebook.com (14 rows),
# youtube.com (10), reddit.com (5), instagram.com (4) and wikipedia (4) ended
# up in a report that goes to the sales team.

# Domains that never contain a usable project lead.
BLOCKED_DOMAINS = {
    # Social / UGC
    "facebook.com", "m.facebook.com", "instagram.com", "twitter.com", "x.com",
    "reddit.com", "youtube.com", "youtu.be", "tiktok.com", "pinterest.com",
    "linkedin.com", "quora.com", "medium.com", "threads.net", "vk.com",
    # Reference
    "wikipedia.org", "wikiwand.com", "fandom.com", "britannica.com",
    # Market-research / lead-gen spam (sells reports, never announces projects)
    "marketdataforecast.com", "mordorintelligence.com", "marketsandmarkets.com",
    "grandviewresearch.com", "researchandmarkets.com", "globenewswire.com",
    "prnewswire.com", "openpr.com", "einpresswire.com", "marketwatch.com",
    "futuremarketinsights.com", "fortunebusinessinsights.com",
    "alliedmarketresearch.com", "imarcgroup.com", "technavio.com",
    "precedenceresearch.com", "verifiedmarketresearch.com",
    "expertmarketresearch.com", "marketresearchfuture.com",
    # Job boards / classifieds / aggregators
    "indeed.com", "glassdoor.com", "ziprecruiter.com", "monster.com",
    "controller.com", "aircraftforsale.com", "trade-a-plane.com",
    "globalair.com", "airport-suppliers.com", "tendersinfo.com",
    "bidnetdirect.com", "biddingo.com",
}

# Marketing / brochure pages. A URL path containing these is almost always a
# vendor's own product page rather than news about a project.
VENDOR_PATH_MARKERS = (
    "/products/", "/product/", "/services/", "/industries/", "/capabilities/",
    "/our-work/", "/portfolio/", "/solutions/", "/shop/", "/store/",
    "/category/", "/catalog/", "/brochure", "/datasheet", "/contact",
    "/about-us", "/about/", "/careers", "/jobs/", "/market-report",
    "/market-reports/", "-market-size", "-market-share", "/wp-content/",
)

# Titles that signal a brochure or an evergreen page, not news.
VENDOR_TITLE_PATTERNS = [
    r"\bfor sale\b", r"\bbuy\b", r"\bshop\b", r"\bprice list\b",
    r"\bmarket (size|share|report|outlook|forecast|analysis)\b",
    r"\bcagr\b", r"\bindustry report\b",
    r"\b(we|our company) (design|build|manufacture|provide)\b",
    r"\bcontact us\b", r"\bhome ?[-|]", r"\bwikipedia\b",
    r"^\s*(airports?|aviation|hangars?)\s*$",     # bare nav-page titles
    r"\bsuppliers?\b.*\bdirectory\b", r"\bdirectory\b.*\bsuppliers?\b",
]

# Words that mean a real project is happening. At least one is required.
PROJECT_SIGNALS = {
    # Strong — money is committed or ground is moving
    "groundbreaking": 6, "broke ground": 6, "breaks ground": 6,
    "contract awarded": 6, "awarded a contract": 6, "awarded the contract": 6,
    "wins contract": 5, "awarded": 4, "construction begins": 6,
    "begins construction": 6, "under construction": 4, "topped out": 5,
    "ribbon cutting": 5, "opens new": 4, "completed": 3,
    # Pipeline — the early-design stage that makes a sales lead
    "approved": 4, "green light": 4, "greenlight": 4, "planning approval": 5,
    "planning permission": 5, "proposed": 3, "plans to build": 5,
    "will build": 4, "to construct": 4, "design phase": 5,
    "feasibility": 3, "request for proposal": 5, "rfp": 4,
    "invitation to tender": 5, "solicitation": 4, "bid": 2,
    "master plan": 3, "environmental assessment": 3,
    # Investment language
    "million": 3, "billion": 4, "investment": 2, "funding": 2,
    "budget": 2, "capital": 1,
    # Facility nouns
    "hangar": 5, "hangars": 5, "mro": 3, "fbo": 3, "maintenance facility": 3,
    "aircraft maintenance": 3, "air base": 2, "airbase": 2, "airfield": 2,
    # Fire protection — highest value to Safespill
    "fire suppression": 6, "fire protection": 5, "foam system": 6,
    "afff": 6, "fluorine-free": 6, "fluorine free": 6, "nfpa 409": 8,
    "f3": 2, "deluge": 5, "sprinkler": 3,
}

# A row must mention at least one of these to be about aviation facilities.
AVIATION_ANCHORS = (
    "hangar", "hangars", "aircraft", "airport", "airfield", "aviation",
    "mro", "fbo", "air base", "airbase", "aerodrome", "airplane",
    "helicopter", "jet ", "airline", "air force", "naval air",
)

MIN_RELEVANCE_SCORE = int(env("MIN_RELEVANCE_SCORE", "12"))


def normalise_url(url: str) -> str:
    """
    Canonical form for dedupe: lowercase host, no www., no query string, no
    fragment, no trailing slash. v1 compared raw URLs, so the same article
    with a ?utm_source= tag counted as new every week.
    """
    if not url:
        return ""
    try:
        p = urlparse(url.strip())
        host = (p.netloc or "").lower()
        if host.startswith("www."):
            host = host[4:]
        path = (p.path or "").rstrip("/")
        return urlunparse(("https", host, path, "", "", "")).lower()
    except Exception:
        return url.strip().lower()


def registrable_domain(url: str) -> str:
    """Return 'example.co.uk' style domain from a URL (best effort)."""
    try:
        host = (urlparse(url).netloc or "").lower()
    except Exception:
        return ""
    if host.startswith("www."):
        host = host[4:]
    return host


def is_blocked_domain(url: str) -> bool:
    host = registrable_domain(url)
    if not host:
        return True
    for bad in BLOCKED_DOMAINS:
        if host == bad or host.endswith("." + bad):
            return True
    return False


def looks_like_vendor_page(url: str, title: str) -> bool:
    path = ""
    try:
        path = (urlparse(url).path or "").lower()
    except Exception:
        pass
    if any(marker in path for marker in VENDOR_PATH_MARKERS):
        return True
    tl = (title or "").lower()
    for pat in VENDOR_TITLE_PATTERNS:
        if re.search(pat, tl):
            return True
    return False


def relevance_score(row: dict) -> int:
    """
    Score a candidate row. Title matches count double. Returns 0 if the row
    has no aviation anchor at all.
    """
    title = (row.get("Project Title") or "").lower()
    summary = (row.get("Summary") or "").lower()
    blob = title + " " + summary

    if not any(a in blob for a in AVIATION_ANCHORS):
        return 0

    score = 0
    for phrase, weight in PROJECT_SIGNALS.items():
        if phrase in title:
            score += weight * 2
        elif phrase in summary:
            score += weight

    # A dollar/currency figure is a strong sign of a real, sized project.
    if re.search(r'[$€£¥]\s?\d|\b\d+(\.\d+)?\s?(million|billion|bn|m)\b', blob):
        score += 5

    # Square footage / area — same idea.
    if re.search(r'\b\d[\d,]*\s?(square (feet|foot|metre|meter)|sq\.?\s?(ft|m)|sqft|m2)\b', blob):
        score += 4

    # Procurement rows are inherently actionable.
    if row.get("Source") in ("sam_gov", "canadabuys", "austender", "ted_europa", "nz_gets"):
        score += 10

    return score


# ==============================================================================
# GEOGRAPHY  —  rewritten for v2
# ==============================================================================
#
# The 8/3 report had a blank Location on 187 of 341 rows (55%) and several
# outright wrong ones (a US ENR MILCON article tagged China; a South Australia
# project routed to the North America tab; US-state "Georgia" articles routed
# to EMEA as the country Georgia).
#
# Root causes in v1 and the v2 fix for each:
#
#   v1 fed the first 3,000 characters of tag-stripped HTML into the detector.
#      On most news sites that is JavaScript, CSS, cookie notices and nav.
#   -> v2 removes <script>/<style>/<nav>/<header>/<footer>/<aside> blocks
#      entirely, then prefers og:description and the first real <p> tags.
#
#   v1 treated every signal as interchangeable and summed raw counts.
#   -> v2 checks the TITLE first and returns immediately on an unambiguous
#      title hit. Trevor's note: the headline usually names the location.
#
#   v1 had "Georgia" in both US_STATES_FULL and EUROPE_COUNTRIES, and
#      "Washington" as a state only.
#   -> v2 routes every ambiguous name through disambiguate_* helpers.
#
#   v1 ignored the publisher entirely.
#   -> v2 uses ccTLD (.co.uk, .com.au, .co.za, .gc.ca) and a domain->country
#      map built from the EMEA_NEWS_DOMAINS list that v1 defined and never used.
#
#   v1 had no airport/air-base knowledge, which is most of this domain's text.
#   -> v2 adds an airport and air-base gazetteer.
#
#   v1 returned "" and classify_region("") defaulted to NA.
#   -> v2 returns "Unknown" and those rows get their own tab.

UNKNOWN = "Unknown"

US_STATES = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT",
    "Delaware": "DE", "Florida": "FL", "Georgia": "GA", "Hawaii": "HI",
    "Idaho": "ID", "Illinois": "IL", "Indiana": "IN", "Iowa": "IA",
    "Kansas": "KS", "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME",
    "Maryland": "MD", "Massachusetts": "MA", "Michigan": "MI",
    "Minnesota": "MN", "Mississippi": "MS", "Missouri": "MO",
    "Montana": "MT", "Nebraska": "NE", "Nevada": "NV",
    "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM",
    "New York": "NY", "North Carolina": "NC", "North Dakota": "ND",
    "Ohio": "OH", "Oklahoma": "OK", "Oregon": "OR", "Pennsylvania": "PA",
    "Rhode Island": "RI", "South Carolina": "SC", "South Dakota": "SD",
    "Tennessee": "TN", "Texas": "TX", "Utah": "UT", "Vermont": "VT",
    "Virginia": "VA", "Washington": "WA", "West Virginia": "WV",
    "Wisconsin": "WI", "Wyoming": "WY",
    "District of Columbia": "DC",
}
US_ABBR_TO_STATE = {v: k for k, v in US_STATES.items()}

CA_PROVINCES = {
    "Alberta": "AB", "British Columbia": "BC", "Manitoba": "MB",
    "New Brunswick": "NB", "Newfoundland and Labrador": "NL",
    "Nova Scotia": "NS", "Ontario": "ON", "Prince Edward Island": "PE",
    "Quebec": "QC", "Québec": "QC", "Saskatchewan": "SK",
    "Northwest Territories": "NT", "Nunavut": "NU", "Yukon": "YT",
}
CA_ABBR_TO_PROVINCE = {
    "AB": "Alberta", "BC": "British Columbia", "MB": "Manitoba",
    "NB": "New Brunswick", "NL": "Newfoundland and Labrador",
    "NS": "Nova Scotia", "ON": "Ontario", "PE": "Prince Edward Island",
    "QC": "Quebec", "SK": "Saskatchewan", "NT": "Northwest Territories",
    "NU": "Nunavut", "YT": "Yukon",
}

AU_STATES = {
    "New South Wales": "NSW", "Victoria": "VIC", "Queensland": "QLD",
    "South Australia": "SA", "Western Australia": "WA", "Tasmania": "TAS",
    "Northern Territory": "NT", "Australian Capital Territory": "ACT",
}

NZ_REGIONS = {
    "Auckland", "Wellington", "Canterbury", "Otago", "Waikato",
    "Bay of Plenty", "Manawatu", "Northland", "Southland", "Taranaki",
    "Hawke's Bay", "Marlborough", "Nelson", "Tasman", "West Coast",
    "Gisborne",
}

MX_STATES = {
    "Aguascalientes", "Baja California", "Campeche", "Chiapas", "Chihuahua",
    "Coahuila", "Colima", "Durango", "Guanajuato", "Guerrero", "Hidalgo",
    "Jalisco", "Mexico City", "Michoacan", "Morelos", "Nayarit",
    "Nuevo Leon", "Oaxaca", "Puebla", "Queretaro", "Quintana Roo",
    "San Luis Potosi", "Sinaloa", "Sonora", "Tabasco", "Tamaulipas",
    "Tlaxcala", "Veracruz", "Yucatan", "Zacatecas",
}

# --- Country -> region --------------------------------------------------------
# Regions: USA, CANADA, ANZ, EMEA. Every country belongs to exactly one.

EUROPE_COUNTRIES = {
    "Albania", "Andorra", "Armenia", "Austria", "Azerbaijan", "Belarus",
    "Belgium", "Bosnia and Herzegovina", "Bulgaria", "Croatia", "Cyprus",
    "Czechia", "Denmark", "Estonia", "Finland", "France", "Georgia",
    "Germany", "Greece", "Hungary", "Iceland", "Ireland", "Italy",
    "Kazakhstan", "Kosovo", "Latvia", "Liechtenstein", "Lithuania",
    "Luxembourg", "Malta", "Moldova", "Monaco", "Montenegro", "Netherlands",
    "North Macedonia", "Norway", "Poland", "Portugal", "Romania", "Russia",
    "San Marino", "Serbia", "Slovakia", "Slovenia", "Spain", "Sweden",
    "Switzerland", "Turkey", "Ukraine", "United Kingdom", "Greenland",
}

MIDDLE_EAST_COUNTRIES = {
    "Bahrain", "Iran", "Iraq", "Israel", "Jordan", "Kuwait", "Lebanon",
    "Oman", "Palestine", "Qatar", "Saudi Arabia", "Syria",
    "United Arab Emirates", "Yemen",
}

AFRICA_COUNTRIES = {
    "Algeria", "Angola", "Benin", "Botswana", "Burkina Faso", "Burundi",
    "Cameroon", "Cape Verde", "Central African Republic", "Chad", "Comoros",
    "Congo", "Democratic Republic of the Congo", "Ivory Coast", "Djibouti",
    "Egypt", "Equatorial Guinea", "Eritrea", "Eswatini", "Ethiopia",
    "Gabon", "Gambia", "Ghana", "Guinea", "Guinea-Bissau", "Kenya",
    "Lesotho", "Liberia", "Libya", "Madagascar", "Malawi", "Mali",
    "Mauritania", "Mauritius", "Morocco", "Mozambique", "Namibia", "Niger",
    "Nigeria", "Rwanda", "Senegal", "Seychelles", "Sierra Leone", "Somalia",
    "South Africa", "South Sudan", "Sudan", "Tanzania", "Togo", "Tunisia",
    "Uganda", "Zambia", "Zimbabwe",
}

# Asia and the rest of the world fold into EMEA — the report has no Asia tab.
# ANZ is now its own tab, so Australia and New Zealand are pulled out.
ASIA_OTHER_COUNTRIES = {
    "Afghanistan", "Bangladesh", "Bhutan", "Brunei", "Cambodia", "China",
    "Hong Kong", "India", "Indonesia", "Japan", "Laos", "Malaysia",
    "Maldives", "Mongolia", "Myanmar", "Nepal", "North Korea", "Pakistan",
    "Philippines", "Singapore", "South Korea", "Sri Lanka", "Taiwan",
    "Thailand", "Timor-Leste", "Uzbekistan", "Vietnam",
    "Argentina", "Brazil", "Chile", "Colombia", "Peru", "Guam",
}

ANZ_COUNTRIES = {"Australia", "New Zealand", "Papua New Guinea", "Fiji"}

COUNTRY_REGION: dict[str, str] = {}
for _c in EUROPE_COUNTRIES | MIDDLE_EAST_COUNTRIES | AFRICA_COUNTRIES | ASIA_OTHER_COUNTRIES:
    COUNTRY_REGION[_c] = "EMEA"
for _c in ANZ_COUNTRIES:
    COUNTRY_REGION[_c] = "ANZ"
COUNTRY_REGION["United States"] = "USA"
COUNTRY_REGION["Canada"] = "CANADA"
COUNTRY_REGION["Mexico"] = "EMEA"   # no LATAM tab; Mexico is rare in this feed

ALL_COUNTRIES = set(COUNTRY_REGION)

# Names that are NOT safe to match bare, because they collide with a US state,
# a US city, or an English word. Each needs corroborating evidence.
AMBIGUOUS_COUNTRY_NAMES = {
    "Georgia",      # US state (Atlanta) vs. the country (Tbilisi)
    "Jordan",       # person's name
    "Chad", "Niger", "Mali", "Congo", "Guinea", "Togo", "Benin",
    "China",        # "China Lake" naval air station is in California
    "India",        # "India" the NATO phonetic letter; Indiana substrings
    "Turkey",       # the bird
    "Israel",       # person's name
    "Oman",         # substring risk
    "Malta",        # Malta, New York / Malta, Illinois
    "Lebanon",      # Lebanon, TN / Lebanon, OH / Lebanon, PA
    "Canada",       # handled explicitly
    "Mexico",       # New Mexico, Mexico MO
}

COUNTRY_ALIASES = {
    "usa": "United States", "u.s.": "United States", "u.s.a.": "United States",
    "united states of america": "United States", "us military": "United States",
    "uk": "United Kingdom", "u.k.": "United Kingdom",
    "britain": "United Kingdom", "great britain": "United Kingdom",
    "england": "United Kingdom", "scotland": "United Kingdom",
    "wales": "United Kingdom", "northern ireland": "United Kingdom",
    "uae": "United Arab Emirates", "u.a.e.": "United Arab Emirates",
    "emirati": "United Arab Emirates",
    "holland": "Netherlands", "the netherlands": "Netherlands",
    "dutch": "Netherlands",
    "czech republic": "Czechia", "saudi": "Saudi Arabia",
    "korea": "South Korea", "south korean": "South Korea",
    "kingdom of saudi arabia": "Saudi Arabia", "ksa": "Saudi Arabia",
    "türkiye": "Turkey", "turkiye": "Turkey",
    "aotearoa": "New Zealand", "kiwi": "New Zealand",
    "aussie": "Australia", "australian": "Australia",
    "german": "Germany", "french": "France", "spanish": "Spain",
    "italian": "Italy", "polish": "Poland", "nigerian": "Nigeria",
    "kenyan": "Kenya", "south african": "South Africa", "egyptian": "Egypt",
    "canadian": "Canada", "british": "United Kingdom", "irish": "Ireland",
}

# ccTLD -> country. Used as evidence about the publisher, weighted modestly:
# a UK trade magazine can report on a US project.
TLD_COUNTRY = {
    ".co.uk": "United Kingdom", ".org.uk": "United Kingdom",
    ".gov.uk": "United Kingdom", ".ac.uk": "United Kingdom",
    ".uk": "United Kingdom",
    ".com.au": "Australia", ".gov.au": "Australia", ".org.au": "Australia",
    ".net.au": "Australia", ".edu.au": "Australia",
    ".co.nz": "New Zealand", ".govt.nz": "New Zealand", ".org.nz": "New Zealand",
    ".nz": "New Zealand",
    ".gc.ca": "Canada", ".ca": "Canada",
    ".co.za": "South Africa", ".org.za": "South Africa", ".gov.za": "South Africa",
    ".ae": "United Arab Emirates", ".sa": "Saudi Arabia", ".qa": "Qatar",
    ".kw": "Kuwait", ".bh": "Bahrain", ".om": "Oman", ".jo": "Jordan",
    ".il": "Israel", ".eg": "Egypt", ".ma": "Morocco", ".tn": "Tunisia",
    ".ng": "Nigeria", ".ke": "Kenya", ".gh": "Ghana", ".et": "Ethiopia",
    ".de": "Germany", ".fr": "France", ".es": "Spain", ".it": "Italy",
    ".nl": "Netherlands", ".be": "Belgium", ".pl": "Poland", ".cz": "Czechia",
    ".at": "Austria", ".ch": "Switzerland", ".se": "Sweden", ".no": "Norway",
    ".dk": "Denmark", ".fi": "Finland", ".ie": "Ireland", ".pt": "Portugal",
    ".gr": "Greece", ".ro": "Romania", ".hu": "Hungary", ".tr": "Turkey",
    ".ua": "Ukraine", ".in": "India", ".sg": "Singapore", ".jp": "Japan",
    ".cn": "China", ".kr": "South Korea", ".my": "Malaysia", ".id": "Indonesia",
    ".th": "Thailand", ".ph": "Philippines", ".mx": "Mexico",
}

# Publisher domain -> country. The EMEA_NEWS_DOMAINS list from v1, finally
# put to use, plus common NA trade press mapped to the United States.
DOMAIN_COUNTRY = {
    "zawya.com": "United Arab Emirates",
    "aviationbusinessme.com": "United Arab Emirates",
    "arabianbusiness.com": "United Arab Emirates",
    "gulfbusiness.com": "United Arab Emirates",
    "mepmiddleeast.com": "United Arab Emirates",
    "agbi.com": "United Arab Emirates",
    "adsadvance.co.uk": "United Kingdom",
    "businessairportinternational.com": "United Kingdom",
    "aircraftinteriorsinternational.com": "United Kingdom",
    "aviationbusinessnews.com": "United Kingdom",
    "engineeringnews.co.za": "South Africa",
    "thisdaylive.com": "Nigeria", "vanguardngr.com": "Nigeria",
    "newtelegraphng.com": "Nigeria", "punchng.com": "Nigeria",
    "thesun.ng": "Nigeria", "legit.ng": "Nigeria", "arise.tv": "Nigeria",
    "thecable.ng": "Nigeria", "businessday.ng": "Nigeria",
    "addisinsight.net": "Ethiopia", "ecofinagency.com": "Africa",
    "apanews.net": "Africa", "united24media.com": "Ukraine",
    "haaretz.com": "Israel", "timesofisrael.com": "Israel",
    "ted.europa.eu": "Europe", "defence-industry.eu": "Europe",
    "asianaviation.com": "Singapore",
    "enr.com": "United States", "constructiondive.com": "United States",
    "bizjournals.com": "United States", "ainonline.com": "United States",
    "aviationpros.com": "United States", "aviationweek.com": "United States",
    "flyingmag.com": "United States", "avweb.com": "United States",
    "defensenews.com": "United States", "breakingdefense.com": "United States",
    "airandspaceforces.com": "United States", "af.mil": "United States",
    "navy.mil": "United States", "defense.gov": "United States",
    "sam.gov": "United States", "usace.army.mil": "United States",
    "canadabuys.canada.ca": "Canada", "tenders.gov.au": "Australia",
    "gets.govt.nz": "New Zealand",
}

# --- Airport / air base gazetteer --------------------------------------------
# The single highest-value addition for this domain. Hangar news is written in
# terms of airports and air bases, not countries. Keys are lowercase and are
# matched as whole phrases.

AIRPORT_LOCATIONS: dict[str, str] = {}


def _reg(location: str, *names: str) -> None:
    for n in names:
        AIRPORT_LOCATIONS[n.lower()] = location


# US Air Force / Navy / Army installations
_reg("United States - Oklahoma", "Tinker Air Force Base", "Tinker AFB", "Altus AFB", "Vance AFB")
_reg("United States - Delaware", "Dover Air Force Base", "Dover AFB")
_reg("United States - California", "Travis Air Force Base", "Travis AFB", "Edwards Air Force Base", "Edwards AFB", "Vandenberg", "Beale AFB", "Lemoore", "Miramar", "China Lake", "Point Mugu", "Moffett Field", "Los Angeles Air Force Base")
_reg("United States - Utah", "Hill Air Force Base", "Hill AFB")
_reg("United States - Georgia", "Robins Air Force Base", "Robins AFB", "Moody Air Force Base", "Moody AFB", "Fort Benning", "Fort Moore", "Hunter Army Airfield", "Dobbins Air Reserve Base")
_reg("United States - Florida", "Tyndall Air Force Base", "Tyndall AFB", "Eglin Air Force Base", "Eglin AFB", "MacDill", "Hurlburt Field", "Patrick Space Force Base", "Jacksonville Naval Air Station", "NAS Jacksonville", "Whiting Field", "NAS Pensacola")
_reg("United States - Nevada", "Nellis Air Force Base", "Nellis AFB", "Creech Air Force Base")
_reg("United States - Ohio", "Wright-Patterson Air Force Base", "Wright-Patterson AFB", "Rickenbacker")
_reg("United States - Nebraska", "Offutt Air Force Base", "Offutt AFB")
_reg("United States - Louisiana", "Barksdale Air Force Base", "Barksdale AFB", "Fort Polk", "Fort Johnson")
_reg("United States - Missouri", "Whiteman Air Force Base", "Whiteman AFB", "Fort Leonard Wood")
_reg("United States - Virginia", "Langley Air Force Base", "Joint Base Langley-Eustis", "Naval Station Norfolk", "NAS Oceana", "Fort Belvoir", "Quantico")
_reg("United States - Maryland", "Joint Base Andrews", "Andrews Air Force Base", "Patuxent River", "NAS Patuxent River", "Aberdeen Proving Ground")
_reg("United States - Illinois", "Scott Air Force Base", "Scott AFB")
_reg("United States - Arizona", "Davis-Monthan", "Luke Air Force Base", "Luke AFB", "Yuma Proving Ground", "MCAS Yuma", "Fort Huachuca")
_reg("United States - South Dakota", "Ellsworth Air Force Base", "Ellsworth AFB")
_reg("United States - North Dakota", "Minot Air Force Base", "Minot AFB", "Grand Forks Air Force Base")
_reg("United States - Washington", "Fairchild Air Force Base", "Fairchild AFB", "Joint Base Lewis-McChord", "McChord Field", "NAS Whidbey Island", "Whidbey Island")
_reg("United States - North Carolina", "Seymour Johnson", "Pope Field", "Fort Bragg", "Fort Liberty", "MCAS Cherry Point", "Cherry Point", "MCAS New River", "Camp Lejeune")
_reg("United States - South Carolina", "Shaw Air Force Base", "Shaw AFB", "Joint Base Charleston", "MCAS Beaufort")
_reg("United States - Arkansas", "Little Rock Air Force Base", "Little Rock AFB", "Ebbing Air National Guard Base")
_reg("United States - Texas", "Sheppard Air Force Base", "Sheppard AFB", "Dyess Air Force Base", "Dyess AFB", "Laughlin Air Force Base", "Randolph Air Force Base", "Joint Base San Antonio", "Lackland", "Kelly Field", "Fort Hood", "Fort Cavazos", "NAS Corpus Christi", "NAS Kingsville", "Goodfellow Air Force Base")
_reg("United States - New Mexico", "Kirtland Air Force Base", "Kirtland AFB", "Holloman Air Force Base", "Holloman AFB", "Cannon Air Force Base", "Cannon AFB", "White Sands")
_reg("United States - Colorado", "Buckley Space Force Base", "Peterson Space Force Base", "Schriever", "Fort Carson", "United States Air Force Academy")
_reg("United States - Montana", "Malmstrom Air Force Base", "Malmstrom AFB")
_reg("United States - Idaho", "Mountain Home Air Force Base", "Mountain Home AFB")
_reg("United States - Mississippi", "Columbus Air Force Base", "Keesler Air Force Base", "Keesler AFB")
_reg("United States - Alabama", "Maxwell Air Force Base", "Maxwell AFB", "Redstone Arsenal", "Fort Rucker", "Fort Novosel")
_reg("United States - Tennessee", "Arnold Air Force Base", "Arnold AFB")
_reg("United States - Massachusetts", "Hanscom Air Force Base", "Hanscom AFB", "Otis Air National Guard Base")
_reg("United States - New Jersey", "Joint Base McGuire-Dix-Lakehurst", "McGuire Air Force Base", "Lakehurst")
_reg("United States - New York", "Niagara Falls Air Reserve Station", "Fort Drum", "Stewart Air National Guard Base")
_reg("United States - Michigan", "Selfridge Air National Guard Base", "Selfridge")
_reg("United States - Alaska", "Joint Base Elmendorf-Richardson", "Elmendorf", "Eielson Air Force Base", "Eielson AFB", "Fort Wainwright")
_reg("United States - Hawaii", "Joint Base Pearl Harbor-Hickam", "Hickam", "MCBH Kaneohe Bay", "Wheeler Army Airfield")
_reg("United States - Kansas", "McConnell Air Force Base", "McConnell AFB", "Fort Riley", "Fort Leavenworth")
_reg("United States - Wyoming", "F.E. Warren Air Force Base", "Francis E. Warren")
_reg("United States - Indiana", "Grissom Air Reserve Base")
_reg("United States - Wisconsin", "Volk Field", "Truax Field")
_reg("United States - Minnesota", "Duluth Air National Guard Base")
_reg("United States - Maine", "Bangor Air National Guard Base", "Brunswick Landing")
_reg("United States - Pennsylvania", "Letterkenny", "Willow Grove", "Fort Indiantown Gap")
_reg("United States - Oregon", "Kingsley Field", "Portland Air National Guard Base")
_reg("United States - Nevada", "Fallon", "NAS Fallon")

# Overseas US installations — these belong to the HOST country's tab
_reg("Japan", "Kadena Air Base", "Kadena", "Yokota Air Base", "Yokota", "Misawa Air Base", "Misawa", "Iwakuni", "Atsugi", "Okinawa", "Futenma")
_reg("South Korea", "Osan Air Base", "Osan", "Kunsan Air Base", "Kunsan", "Camp Humphreys")
_reg("Germany", "Ramstein Air Base", "Ramstein", "Spangdahlem", "Wiesbaden Army Airfield", "Geilenkirchen")
_reg("Italy", "Aviano Air Base", "Aviano", "Sigonella", "Naval Air Station Sigonella")
_reg("United Kingdom", "RAF Lakenheath", "Lakenheath", "RAF Mildenhall", "Mildenhall", "RAF Fairford", "Fairford", "RAF Marham", "Marham", "RAF Coningsby", "Coningsby", "RAF Brize Norton", "Brize Norton", "RAF Waddington", "Waddington", "RAF Lossiemouth", "Lossiemouth", "RAF Valley", "RAF Leeming", "RAF Wittering", "RAF Cranwell", "RAF Northolt", "RAF Benson", "RAF Odiham")
_reg("Turkey", "Incirlik Air Base", "Incirlik")
_reg("Qatar", "Al Udeid Air Base", "Al Udeid")
_reg("United Arab Emirates", "Al Dhafra Air Base", "Al Dhafra")
_reg("Kuwait", "Ali Al Salem", "Camp Arifjan")
_reg("Bahrain", "Isa Air Base", "NSA Bahrain")
_reg("Jordan", "Muwaffaq Salti", "Azraq")
_reg("Poland", "Powidz", "Lask Air Base", "Łask", "Redzikowo", "Poznan-Krzesiny", "Krzesiny")
_reg("Romania", "Mihail Kogalniceanu", "Campia Turzii")
_reg("Bulgaria", "Graf Ignatievo", "Bezmer")
_reg("Estonia", "Amari Air Base", "Ämari")
_reg("Lithuania", "Siauliai", "Šiauliai")
_reg("Spain", "Moron Air Base", "Rota Naval Station", "Naval Station Rota", "Torrejon")
_reg("Portugal", "Lajes Field", "Lajes")
_reg("Greenland", "Pituffik", "Thule Air Base")
_reg("Guam", "Andersen Air Force Base", "Andersen AFB")
_reg("Netherlands", "Volkel Air Base", "Leeuwarden Air Base", "Woensdrecht")
_reg("Belgium", "Kleine Brogel", "Florennes")
_reg("Norway", "Orland Air Station", "Ørland", "Rygge")
_reg("Denmark", "Skrydstrup")
_reg("Australia", "RAAF Base Amberley", "Amberley", "RAAF Base Williamtown", "Williamtown", "RAAF Base Tindal", "Tindal", "RAAF Base Edinburgh", "RAAF Base Richmond", "RAAF Base Pearce", "RAAF Base Townsville", "RAAF Base Darwin", "RAAF Base Wagga", "RAAF Base East Sale")
_reg("New Zealand", "RNZAF Base Ohakea", "Ohakea", "RNZAF Base Woodbourne", "Woodbourne", "RNZAF Base Whenuapai", "Whenuapai")
_reg("Canada", "CFB Trenton", "8 Wing Trenton", "CFB Cold Lake", "4 Wing Cold Lake", "CFB Bagotville", "3 Wing Bagotville", "CFB Comox", "19 Wing Comox", "CFB Greenwood", "14 Wing Greenwood", "CFB Goose Bay", "5 Wing Goose Bay", "CFB Borden", "CFB Gagetown")
_reg("Israel", "Nevatim", "Ramat David", "Hatzerim")
_reg("Saudi Arabia", "Prince Sultan Air Base", "King Abdulaziz Air Base")
_reg("Egypt", "Cairo West Air Base")
_reg("India", "Hindon Air Force Station", "Hindan")

# Major civil airports (a subset that actually shows up in hangar news)
_reg("United States - Texas", "Dallas Fort Worth International", "DFW Airport", "Dallas Love Field", "Love Field", "George Bush Intercontinental", "Houston Hobby", "Conroe North Houston Regional", "Austin-Bergstrom", "San Antonio International", "Alliance Airport", "Addison Airport", "Sugar Land Regional", "Ellington Field")
_reg("United States - Georgia", "Hartsfield-Jackson", "Atlanta International", "DeKalb-Peachtree", "Savannah/Hilton Head")
_reg("United States - Colorado", "Denver International", "Rocky Mountain Metropolitan", "Centennial Airport")
_reg("United States - Illinois", "O'Hare International", "Chicago Midway", "Chicago Executive")
_reg("United States - California", "Los Angeles International", "San Francisco International", "Van Nuys Airport", "Long Beach Airport", "John Wayne Airport", "Hollywood Burbank", "Oakland International", "San Diego International", "Sacramento International", "Palm Springs International", "Camarillo Airport", "Hayward Executive")
_reg("United States - Florida", "Miami International", "Orlando International", "Fort Lauderdale-Hollywood", "Tampa International", "Jacksonville International", "Opa-locka", "Boca Raton Airport", "Naples Airport", "Sarasota Bradenton", "Daytona Beach International", "Palm Beach International")
_reg("United States - New York", "John F. Kennedy International", "JFK Airport", "LaGuardia", "Republic Airport", "Westchester County Airport", "Stewart International")
_reg("United States - New Jersey", "Newark Liberty", "Teterboro Airport", "Teterboro", "Atlantic City International", "Morristown Municipal")
_reg("United States - Arizona", "Phoenix Sky Harbor", "Scottsdale Airport", "Phoenix-Mesa Gateway", "Tucson International", "Pinal Airpark")
_reg("United States - Nevada", "Harry Reid International", "McCarran International", "Reno-Tahoe International", "Henderson Executive")
_reg("United States - Washington", "Seattle-Tacoma International", "Paine Field", "Boeing Field", "King County International", "Spokane International")
_reg("United States - Oregon", "Portland International", "Hillsboro Airport")
_reg("United States - Utah", "Salt Lake City International", "Provo Airport")
_reg("United States - Massachusetts", "Logan International", "Hanscom Field", "Worcester Regional")
_reg("United States - Michigan", "Detroit Metropolitan", "Willow Run", "Oakland County International")
_reg("United States - Ohio", "Cleveland Hopkins", "John Glenn Columbus", "Cincinnati/Northern Kentucky", "Rickenbacker International")
_reg("United States - North Carolina", "Charlotte Douglas", "Raleigh-Durham International", "Piedmont Triad", "Wilmington International", "Concord-Padgett")
_reg("United States - Tennessee", "Nashville International", "Memphis International", "Chattanooga Metropolitan", "McGhee Tyson")
_reg("United States - Minnesota", "Minneapolis-Saint Paul International", "Flying Cloud", "Anoka County")
_reg("United States - Missouri", "Kansas City International", "St. Louis Lambert", "Spirit of St. Louis")
_reg("United States - Indiana", "Indianapolis International", "Fort Wayne International")
_reg("United States - Kansas", "Wichita Dwight D. Eisenhower", "Wichita Eisenhower")
_reg("United States - Pennsylvania", "Philadelphia International", "Pittsburgh International", "Lehigh Valley International")
_reg("United States - Virginia", "Washington Dulles", "Dulles International", "Ronald Reagan Washington National", "Richmond International", "Manassas Regional", "Leesburg Executive")
_reg("United States - Maryland", "Baltimore/Washington International", "BWI Marshall", "Martin State Airport")
_reg("United States - Connecticut", "Bradley International", "Sikorsky Memorial")
_reg("United States - Wisconsin", "Milwaukee Mitchell", "Appleton International", "Wittman Regional")
_reg("United States - Iowa", "Des Moines International", "Eastern Iowa Airport")
_reg("United States - Nebraska", "Eppley Airfield")
_reg("United States - Oklahoma", "Will Rogers World", "Tulsa International")
_reg("United States - Arkansas", "Bill and Hillary Clinton National", "Northwest Arkansas National")
_reg("United States - Louisiana", "Louis Armstrong New Orleans", "Shreveport Regional")
_reg("United States - Alabama", "Birmingham-Shuttlesworth", "Huntsville International")
_reg("United States - Mississippi", "Jackson-Medgar Wiley Evers")
_reg("United States - Kentucky", "Louisville Muhammad Ali", "Blue Grass Airport", "Cincinnati Northern Kentucky")
_reg("United States - South Carolina", "Charleston International", "Greenville-Spartanburg", "Columbia Metropolitan", "Myrtle Beach International")
_reg("United States - New Mexico", "Albuquerque International Sunport", "Roswell Air Center")
_reg("United States - Idaho", "Boise Airport")
_reg("United States - Montana", "Bozeman Yellowstone", "Billings Logan")
_reg("United States - Alaska", "Ted Stevens Anchorage", "Fairbanks International")
_reg("United States - Hawaii", "Daniel K. Inouye International", "Honolulu International", "Kahului Airport")
_reg("United States - Vermont", "Burlington International")
_reg("United States - Maine", "Portland International Jetport", "Bangor International")
_reg("United States - West Virginia", "Yeager Airport")
_reg("United States - Delaware", "Wilmington Airport")
_reg("United States - Rhode Island", "T.F. Green", "Rhode Island T.F. Green")
_reg("United States - North Dakota", "Hector International")
_reg("United States - South Dakota", "Sioux Falls Regional")
_reg("United States - Wyoming", "Jackson Hole Airport")
_reg("United States - New Hampshire", "Manchester-Boston Regional", "Portsmouth International")
_reg("United States - Nevada", "Reno Stead")

_reg("Canada - Ontario", "Toronto Pearson", "Pearson International", "Billy Bishop", "Ottawa Macdonald-Cartier", "London International Airport", "Region of Waterloo International", "Hamilton International", "John C. Munro")
_reg("Canada - Quebec", "Montreal-Trudeau", "Pierre Elliott Trudeau", "Montreal Mirabel", "Mirabel", "Quebec City Jean Lesage", "Saint-Hubert")
_reg("Canada - British Columbia", "Vancouver International", "Victoria International", "Kelowna International", "Abbotsford International")
_reg("Canada - Alberta", "Calgary International", "Edmonton International", "Springbank Airport")
_reg("Canada - Manitoba", "Winnipeg Richardson", "Richardson International")
_reg("Canada - Saskatchewan", "Saskatoon John G. Diefenbaker", "Regina International")
_reg("Canada - Nova Scotia", "Halifax Stanfield")
_reg("Canada - New Brunswick", "Moncton Roméo LeBlanc", "Fredericton International", "Saint John Airport")
_reg("Canada - Newfoundland and Labrador", "St. John's International", "Gander International")

_reg("United Arab Emirates", "Dubai International", "Al Maktoum International", "Dubai World Central", "Abu Dhabi International", "Zayed International", "Sharjah International", "Al Bateen Executive", "Dubai South")
_reg("Saudi Arabia", "King Khalid International", "King Abdulaziz International", "King Fahd International", "King Salman International", "Red Sea International Airport")
_reg("Qatar", "Hamad International", "Doha International")
_reg("Kuwait", "Kuwait International")
_reg("Bahrain", "Bahrain International")
_reg("Oman", "Muscat International", "Salalah Airport")
_reg("Jordan", "Queen Alia International")
_reg("Israel", "Ben Gurion Airport", "Ben Gurion International")
_reg("Egypt", "Cairo International", "Sphinx International", "Borg El Arab")
_reg("Nigeria", "Murtala Muhammed International", "Nnamdi Azikiwe International", "Lagos airport", "Abuja airport")
_reg("Kenya", "Jomo Kenyatta International", "Moi International")
_reg("South Africa", "O.R. Tambo International", "OR Tambo", "Cape Town International", "King Shaka International", "Lanseria International", "Wonderboom")
_reg("Ethiopia", "Bole International", "Addis Ababa Bole")
_reg("Morocco", "Mohammed V International", "Marrakesh Menara")
_reg("Ghana", "Kotoka International")
_reg("Rwanda", "Kigali International", "Bugesera International")
_reg("Tanzania", "Julius Nyerere International")
_reg("Senegal", "Blaise Diagne International")

_reg("United Kingdom", "Heathrow", "Gatwick", "Stansted", "Luton Airport", "Manchester Airport", "Birmingham Airport", "Edinburgh Airport", "Glasgow Airport", "Bristol Airport", "Newcastle Airport", "Leeds Bradford", "Farnborough Airport", "Biggin Hill", "London City Airport", "Doncaster Sheffield", "Prestwick", "Cardiff Airport", "Belfast International", "Norwich Airport", "Cranfield Airport", "Teesside International")
_reg("Ireland", "Dublin Airport", "Shannon Airport", "Cork Airport", "Weston Airport")
_reg("Germany", "Frankfurt Airport", "Munich Airport", "Berlin Brandenburg", "Hamburg Airport", "Dusseldorf Airport", "Düsseldorf Airport", "Cologne Bonn", "Stuttgart Airport", "Leipzig/Halle", "Nuremberg Airport", "Hannover Airport")
_reg("France", "Charles de Gaulle", "Paris Orly", "Le Bourget", "Nice Cote d'Azur", "Lyon Saint-Exupery", "Marseille Provence", "Toulouse-Blagnac", "Bordeaux-Merignac", "Chateauroux")
_reg("Netherlands", "Amsterdam Schiphol", "Schiphol", "Rotterdam The Hague Airport", "Eindhoven Airport", "Maastricht Aachen")
_reg("Spain", "Madrid-Barajas", "Barajas", "Barcelona-El Prat", "El Prat", "Malaga-Costa del Sol", "Palma de Mallorca Airport", "Valencia Airport", "Teruel Airport")
_reg("Italy", "Rome Fiumicino", "Fiumicino", "Milan Malpensa", "Malpensa", "Linate", "Venice Marco Polo", "Naples International", "Bologna Guglielmo Marconi")
_reg("Belgium", "Brussels Airport", "Liege Airport", "Brussels South Charleroi", "Ostend-Bruges")
_reg("Switzerland", "Zurich Airport", "Geneva Airport", "EuroAirport Basel")
_reg("Austria", "Vienna International", "Vienna Schwechat")
_reg("Poland", "Warsaw Chopin", "Chopin Airport", "Katowice Airport", "Krakow Airport", "Gdansk Lech Walesa", "Rzeszow-Jasionka", "Warsaw Modlin")
_reg("Czechia", "Vaclav Havel Airport", "Prague Airport", "Brno-Turany")
_reg("Hungary", "Budapest Ferenc Liszt", "Ferihegy")
_reg("Romania", "Henri Coanda", "Otopeni", "Cluj Avram Iancu")
_reg("Greece", "Athens International", "Eleftherios Venizelos", "Thessaloniki Makedonia")
_reg("Portugal", "Humberto Delgado", "Lisbon Airport", "Porto Airport", "Beja Airport")
_reg("Sweden", "Stockholm Arlanda", "Arlanda", "Gothenburg Landvetter")
_reg("Norway", "Oslo Gardermoen", "Gardermoen", "Bergen Flesland", "Stavanger Sola")
_reg("Denmark", "Copenhagen Airport", "Kastrup", "Billund Airport")
_reg("Finland", "Helsinki-Vantaa", "Vantaa")
_reg("Iceland", "Keflavik International", "Keflavík")
_reg("Turkey", "Istanbul Airport", "Sabiha Gokcen", "Esenboga", "Antalya Airport")
_reg("Ukraine", "Boryspil International", "Boryspil")
_reg("Malta", "Malta International Airport", "Luqa")
_reg("Cyprus", "Larnaca International", "Paphos International")
_reg("Luxembourg", "Findel Airport", "Luxembourg Findel")
_reg("Serbia", "Nikola Tesla Airport")
_reg("Georgia", "Tbilisi International", "Tbilisi")
_reg("Kazakhstan", "Almaty International", "Astana International", "Nursultan Nazarbayev")

_reg("Australia - New South Wales", "Sydney Airport", "Kingsford Smith", "Bankstown Airport", "Western Sydney International", "Newcastle Airport Williamtown", "Wagga Wagga Airport")
_reg("Australia - Victoria", "Melbourne Airport", "Tullamarine", "Avalon Airport", "Essendon Fields", "Moorabbin Airport")
_reg("Australia - Queensland", "Brisbane Airport", "Gold Coast Airport", "Cairns Airport", "Townsville Airport", "Toowoomba Wellcamp", "Archerfield Airport")
_reg("Australia - Western Australia", "Perth Airport", "Jandakot Airport", "Karratha Airport")
_reg("Australia - South Australia", "Adelaide Airport", "Parafield Airport")
_reg("Australia - Tasmania", "Hobart Airport", "Launceston Airport")
_reg("Australia - Northern Territory", "Darwin International", "Alice Springs Airport")
_reg("Australia - Australian Capital Territory", "Canberra Airport")
_reg("New Zealand", "Auckland Airport", "Auckland International", "Christchurch International", "Wellington International", "Queenstown Airport", "Hamilton Airport New Zealand", "Dunedin Airport", "Palmerston North Airport")

_reg("Singapore", "Changi Airport", "Seletar Airport")
_reg("Japan", "Narita International", "Haneda Airport", "Kansai International", "Chubu Centrair")
_reg("South Korea", "Incheon International", "Gimpo International")
_reg("China", "Beijing Capital International", "Beijing Daxing", "Shanghai Pudong", "Guangzhou Baiyun", "Shenzhen Bao'an", "Hong Kong International")
_reg("India", "Indira Gandhi International", "Chhatrapati Shivaji", "Kempegowda International", "Rajiv Gandhi International", "Noida International", "Navi Mumbai International")
_reg("Malaysia", "Kuala Lumpur International", "Subang Airport", "Sultan Abdul Aziz Shah")
_reg("Indonesia", "Soekarno-Hatta")
_reg("Thailand", "Suvarnabhumi", "Don Mueang", "U-Tapao")
_reg("Philippines", "Ninoy Aquino International", "Clark International")
_reg("Vietnam", "Tan Son Nhat", "Noi Bai International", "Long Thanh International")
_reg("Taiwan", "Taoyuan International")
_reg("Brazil", "Guarulhos", "Galeao", "Viracopos")
_reg("Mexico", "Benito Juarez International", "Felipe Angeles International", "Monterrey International", "Queretaro Intercontinental")
_reg("United States - Washington", "Sea-Tac", "SeaTac", "Seatac")

# Coast Guard air stations and Naval Air Stations — these turn up constantly in
# hangar procurement news and v1 knew nothing about them.
_reg("United States - Hawaii", "Barbers Point", "Air Station Barbers Point")
_reg("United States - North Carolina", "Air Station Elizabeth City", "Elizabeth City")
_reg("United States - Alabama", "Aviation Training Center Mobile", "Air Station Mobile")
_reg("United States - Florida", "Air Station Clearwater", "Air Station Miami", "Air Station Savannah")
_reg("United States - Michigan", "Air Station Traverse City", "Traverse City")
_reg("United States - Oregon", "Air Station North Bend")
_reg("United States - California", "Air Station Sacramento", "Air Station San Francisco", "Air Station Humboldt Bay", "Air Station Los Angeles", "Naval Air Weapons Station")
_reg("United States - Texas", "Air Station Corpus Christi", "Air Station Houston")
_reg("United States - Alaska", "Air Station Kodiak", "Air Station Sitka", "Kodiak")
_reg("United States - New Jersey", "Naval Air Station Wildwood", "Air Station Atlantic City", "Wildwood")
_reg("United States - Massachusetts", "Air Station Cape Cod")
_reg("United States - Louisiana", "Air Station New Orleans")
_reg("United States - Puerto Rico", "Air Station Borinquen")

# --- US cities ----------------------------------------------------------------
#
# The single biggest remaining gap after the airport gazetteer. US aviation
# news names the city, not the state: "Million Air breaks ground on major
# AUSTIN FBO expansion", "a legacy LONG BEACH hangar". v1 had no city list at
# all, so all of those rows came back with a blank Location.
#
# Deliberately EXCLUDED: city names that collide with a non-US city already in
# COUNTRY_CITIES (Birmingham, Manchester, Newcastle, Cambridge, Athens, Rome,
# Paris, Berlin, Moscow, Toledo, Naples, Venice, Vienna, Odessa, Dublin,
# Glasgow, Bristol, Hamburg, Brighton, Coventry) or across several US states
# (Springfield, Columbus, Portland, Charleston, Arlington, Aurora, Glendale,
# Pasadena, Peoria, Rochester, Richmond, Jackson, Columbia, Lafayette).
# Those rely on the state name or the "City, ST" pattern instead.

US_CITIES = {
    "Alabama": ["Huntsville", "Montgomery", "Tuscaloosa", "Dothan", "Auburn"],
    "Alaska": ["Anchorage", "Fairbanks", "Juneau", "Ketchikan"],
    "Arizona": ["Phoenix", "Tucson", "Mesa", "Chandler", "Scottsdale", "Tempe", "Prescott", "Marana"],
    "Arkansas": ["Little Rock", "Fayetteville", "Fort Smith", "Bentonville"],
    "California": ["Los Angeles", "San Diego", "San Jose", "San Francisco", "Fresno",
                   "Sacramento", "Long Beach", "Oakland", "Bakersfield", "Anaheim",
                   "Santa Ana", "Irvine", "Chula Vista", "Fremont", "Modesto",
                   "Santa Clarita", "Oxnard", "Fontana", "Moreno Valley", "Huntington Beach",
                   "Santa Rosa", "Palmdale", "Salinas", "Hayward", "Escondido",
                   "Sunnyvale", "Torrance", "Pomona", "Palo Alto", "Burbank",
                   "Van Nuys", "Camarillo", "Carlsbad", "Santa Monica", "Mojave",
                   "Victorville", "Redlands", "Chino", "Hawthorne", "El Segundo",
                   "Silicon Valley", "Palm Springs", "Monterey", "Santa Barbara"],
    "Colorado": ["Denver", "Colorado Springs", "Fort Collins", "Boulder", "Broomfield",
                 "Grand Junction", "Pueblo", "Englewood", "Loveland"],
    "Connecticut": ["Hartford", "New Haven", "Stamford", "Bridgeport", "Groton", "Windsor Locks"],
    "Delaware": ["Wilmington", "Newark Delaware", "New Castle County"],
    "District of Columbia": ["Washington D.C.", "Washington DC", "Washington, D.C."],
    "Florida": ["Jacksonville", "Miami", "Tampa", "Orlando", "St. Petersburg",
                "Hialeah", "Tallahassee", "Fort Lauderdale", "Cape Coral",
                "Pembroke Pines", "Hollywood Florida", "Gainesville", "Coral Springs",
                "Clearwater", "Palm Bay", "West Palm Beach", "Pompano Beach",
                "Boca Raton", "Sarasota", "Melbourne Florida", "Daytona Beach",
                "Naples Florida", "Fort Myers", "Pensacola", "Ocala", "Opa-locka",
                "Kissimmee", "Titusville", "Panama City", "Destin", "Sanford"],
    "Georgia": ["Atlanta", "Augusta", "Savannah", "Macon", "Alpharetta", "Marietta",
                "Warner Robins", "Valdosta", "Brunswick Georgia", "Peachtree City"],
    "Hawaii": ["Honolulu", "Pearl Harbor", "Kahului", "Kona", "Hilo"],
    "Idaho": ["Boise", "Idaho Falls", "Coeur d'Alene", "Nampa", "Pocatello"],
    "Illinois": ["Chicago", "Rockford", "Joliet", "Naperville", "Elgin", "Champaign",
                 "Bloomington Illinois", "Decatur", "Schaumburg", "Belleville"],
    "Indiana": ["Indianapolis", "Fort Wayne", "Evansville", "South Bend", "Carmel",
                "Bloomington Indiana", "Terre Haute", "Gary"],
    "Iowa": ["Des Moines", "Cedar Rapids", "Davenport", "Sioux City", "Iowa City", "Ankeny"],
    "Kansas": ["Wichita", "Overland Park", "Topeka", "Olathe", "Salina"],
    "Kentucky": ["Louisville", "Lexington", "Bowling Green", "Owensboro", "Covington", "Erlanger"],
    "Louisiana": ["New Orleans", "Baton Rouge", "Shreveport", "Lake Charles",
                  "Monroe Louisiana", "Bossier City"],
    "Maine": ["Bangor", "Augusta Maine", "Brunswick Maine", "Presque Isle"],
    "Maryland": ["Baltimore", "Annapolis", "Frederick", "Rockville", "Hagerstown",
                 "Silver Spring", "Bethesda", "Patuxent"],
    "Massachusetts": ["Boston", "Worcester", "Springfield Massachusetts", "Lowell",
                      "Bedford Massachusetts", "Quincy", "Beverly", "New Bedford"],
    "Michigan": ["Detroit", "Grand Rapids", "Ann Arbor", "Lansing", "Flint",
                 "Kalamazoo", "Ypsilanti", "Sterling Heights", "Battle Creek", "Oscoda"],
    "Minnesota": ["Minneapolis", "Saint Paul", "St. Paul", "Duluth", "Rochester Minnesota",
                  "Bloomington Minnesota", "Eden Prairie", "Brooklyn Park"],
    "Mississippi": ["Jackson Mississippi", "Gulfport", "Biloxi", "Hattiesburg", "Meridian", "Tupelo"],
    "Missouri": ["Kansas City", "St. Louis", "Saint Louis", "Springfield Missouri",
                 "Columbia Missouri", "Chesterfield", "Branson"],
    "Montana": ["Billings", "Missoula", "Great Falls", "Bozeman", "Helena"],
    "Nebraska": ["Omaha", "Lincoln Nebraska", "Bellevue Nebraska", "Grand Island"],
    "Nevada": ["Las Vegas", "Reno", "Henderson", "North Las Vegas", "Sparks Nevada", "Fallon"],
    "New Hampshire": ["Manchester New Hampshire", "Nashua", "Concord New Hampshire", "Portsmouth New Hampshire"],
    "New Jersey": ["Newark", "Jersey City", "Trenton", "Paterson", "Teterboro",
                   "Morristown", "Atlantic City", "Camden", "Edison", "Lakehurst"],
    "New Mexico": ["Albuquerque", "Santa Fe", "Las Cruces", "Roswell", "Clovis", "Alamogordo"],
    "New York": ["New York City", "Buffalo", "Yonkers", "Syracuse", "Albany",
                 "New Rochelle", "White Plains", "Farmingdale", "Islip", "Elmira",
                 "Binghamton", "Poughkeepsie", "Long Island", "Brooklyn", "Queens",
                 "Manhattan", "Bronx", "Staten Island", "Westchester"],
    "North Carolina": ["Charlotte", "Raleigh", "Greensboro", "Durham", "Winston-Salem",
                       "Fayetteville North Carolina", "Cary", "Wilmington North Carolina",
                       "Asheville", "Concord North Carolina", "Greenville North Carolina", "Kinston"],
    "North Dakota": ["Fargo", "Bismarck", "Grand Forks", "Minot"],
    "Ohio": ["Cleveland", "Cincinnati", "Toledo Ohio", "Akron", "Dayton",
             "Canton", "Youngstown", "Columbus Ohio", "Wilmington Ohio"],
    "Oklahoma": ["Oklahoma City", "Tulsa", "Norman", "Broken Arrow", "Lawton", "Enid", "Stillwater"],
    "Oregon": ["Portland Oregon", "Eugene", "Salem Oregon", "Gresham", "Hillsboro", "Bend Oregon", "Medford"],
    "Pennsylvania": ["Philadelphia", "Pittsburgh", "Allentown", "Erie Pennsylvania",
                     "Harrisburg", "Scranton", "Bethlehem", "Lancaster", "Coatesville"],
    "Rhode Island": ["Providence", "Warwick", "Pawtucket", "Quonset"],
    "South Carolina": ["Charleston South Carolina", "North Charleston", "Greenville South Carolina",
                       "Columbia South Carolina", "Myrtle Beach", "Spartanburg", "Summerville"],
    "South Dakota": ["Sioux Falls", "Rapid City", "Aberdeen South Dakota"],
    "Tennessee": ["Nashville", "Memphis", "Knoxville", "Chattanooga", "Clarksville",
                  "Murfreesboro", "Smyrna", "Tullahoma"],
    "Texas": ["Houston", "San Antonio", "Dallas", "Austin", "Fort Worth", "El Paso",
              "Corpus Christi", "Plano", "Laredo", "Lubbock", "Garland", "Irving",
              "Amarillo", "Grand Prairie", "Brownsville", "McKinney", "Frisco",
              "Killeen", "McAllen", "Waco", "Midland", "Abilene", "Denton",
              "Beaumont", "Round Rock", "Wichita Falls", "College Station",
              "Sugar Land", "Conroe", "Galveston", "Tyler", "San Marcos",
              "New Braunfels", "Addison", "Georgetown Texas", "Del Rio", "Big Spring"],
    "Utah": ["Salt Lake City", "West Valley City", "Provo", "Ogden", "St. George", "Layton"],
    "Vermont": ["Burlington Vermont", "Montpelier", "South Burlington"],
    "Virginia": ["Virginia Beach", "Norfolk", "Chesapeake", "Newport News",
                 "Alexandria Virginia", "Hampton", "Roanoke", "Portsmouth Virginia",
                 "Lynchburg", "Chantilly", "Manassas", "Leesburg", "Herndon", "Reston", "Danville"],
    "Washington": ["Seattle", "Spokane", "Tacoma", "Vancouver Washington", "Bellevue Washington",
                   "Everett", "Kent Washington", "Renton", "Yakima", "Olympia", "Moses Lake"],
    "West Virginia": ["Charleston West Virginia", "Huntington", "Morgantown", "Martinsburg"],
    "Wisconsin": ["Milwaukee", "Madison", "Green Bay", "Kenosha", "Appleton", "Oshkosh", "Janesville"],
    "Wyoming": ["Cheyenne", "Casper", "Jackson Hole", "Laramie"],
}

for _state, _cities in US_CITIES.items():
    _reg("United States - " + _state, *_cities)

# --- Broad regional terms ----------------------------------------------------
# Not a country, but enough to route a row to the right tab instead of
# stranding it in Needs Review.
REGION_TERMS = {
    "west africa": "Africa", "east africa": "Africa", "north africa": "Africa",
    "sub-saharan africa": "Africa", "southern africa": "Africa",
    "middle east": "Middle East", "gulf region": "Middle East",
    "gcc countries": "Middle East", "persian gulf": "Middle East",
    "arabian gulf": "Middle East",
    "southeast asia": "Asia", "south asia": "Asia", "central asia": "Asia",
    "european union": "Europe", "western europe": "Europe",
    "eastern europe": "Europe", "central europe": "Europe", "nordic": "Europe",
    "scandinavia": "Europe", "benelux": "Europe", "baltics": "Europe",
}

# US federal bodies. If one of these appears but no state can be pinned down,
# the row is at least confidently American rather than Unknown.
US_FEDERAL_MARKERS = (
    "u.s. coast guard", "us coast guard", "coast guard", "department of state",
    "department of defense", "u.s. air force", "us air force", "usaf",
    "u.s. navy", "us navy", "u.s. army", "us army", "army corps of engineers",
    "marine corps", "space force", "national guard", "air national guard",
    "federal aviation administration", "faa", "nasa", "pentagon",
    "general services administration", "veterans affairs", "naval facilities",
    "navfac", "afcec", "air force civil engineer center", "defense logistics",
)

# --- Airline names ------------------------------------------------------------
#
# Aviation news is full of carrier names that CONTAIN a country or state:
# "Alaska Airlines", "Air Canada", "China Eastern", "Air India",
# "Qatar Airways", "Turkish Airlines", "South African Airways".
#
# Counting those as location evidence is wrong — an Alaska Airlines hangar is
# usually in Seattle, and an Air Canada hangar could be anywhere. So carrier
# names are MASKED OUT of the text before country/state/city scoring, and the
# airline's home country is added back as a weak, separate signal. That way
# "Air New Zealand builds a hangar" still resolves to New Zealand, while
# "Alaska Airlines breaks ground at Sea-Tac" correctly resolves to Washington.

AIRLINE_HOME = {
    "alaska airlines": "United States", "alaska air": "United States",
    "hawaiian airlines": "United States", "american airlines": "United States",
    "united airlines": "United States", "delta air lines": "United States",
    "southwest airlines": "United States", "jetblue": "United States",
    "frontier airlines": "United States", "spirit airlines": "United States",
    "allegiant air": "United States", "breeze airways": "United States",
    "avelo airlines": "United States", "sun country airlines": "United States",
    "air canada": "Canada", "westjet": "Canada", "porter airlines": "Canada",
    "air transat": "Canada", "flair airlines": "Canada",
    "air france": "France", "lufthansa": "Germany", "eurowings": "Germany",
    "condor airlines": "Germany", "klm": "Netherlands",
    "british airways": "United Kingdom", "easyjet": "United Kingdom",
    "virgin atlantic": "United Kingdom", "jet2": "United Kingdom",
    "iberia": "Spain", "vueling": "Spain", "air europa": "Spain",
    "ita airways": "Italy", "alitalia": "Italy",
    "swiss international": "Switzerland", "austrian airlines": "Austria",
    "brussels airlines": "Belgium", "scandinavian airlines": "Sweden",
    "norwegian air": "Norway", "finnair": "Finland", "icelandair": "Iceland",
    "aer lingus": "Ireland", "ryanair": "Ireland", "wizz air": "Hungary",
    "lot polish airlines": "Poland", "czech airlines": "Czechia",
    "tap air portugal": "Portugal", "aegean airlines": "Greece",
    "turkish airlines": "Turkey", "pegasus airlines": "Turkey",
    "tarom": "Romania", "air serbia": "Serbia", "airbaltic": "Latvia",
    "ukraine international airlines": "Ukraine", "aeroflot": "Russia",
    "qatar airways": "Qatar", "emirates airline": "United Arab Emirates",
    "etihad airways": "United Arab Emirates", "air arabia": "United Arab Emirates",
    "flydubai": "United Arab Emirates", "saudia": "Saudi Arabia",
    "riyadh air": "Saudi Arabia", "flynas": "Saudi Arabia",
    "kuwait airways": "Kuwait", "gulf air": "Bahrain", "oman air": "Oman",
    "royal jordanian": "Jordan", "el al": "Israel",
    "middle east airlines": "Lebanon",
    "egyptair": "Egypt", "royal air maroc": "Morocco", "tunisair": "Tunisia",
    "kenya airways": "Kenya", "ethiopian airlines": "Ethiopia",
    "south african airways": "South Africa", "airlink": "South Africa",
    "air tanzania": "Tanzania", "rwandair": "Rwanda",
    "air peace": "Nigeria", "ibom air": "Nigeria", "arik air": "Nigeria",
    "air senegal": "Senegal",
    "qantas": "Australia", "jetstar": "Australia",
    "virgin australia": "Australia", "rex airlines": "Australia",
    "air new zealand": "New Zealand", "sounds air": "New Zealand",
    "air india": "India", "indigo airlines": "India", "vistara": "India",
    "spicejet": "India", "akasa air": "India",
    "singapore airlines": "Singapore", "scoot airlines": "Singapore",
    "malaysia airlines": "Malaysia", "airasia": "Malaysia",
    "thai airways": "Thailand", "bangkok airways": "Thailand",
    "vietnam airlines": "Vietnam", "vietjet": "Vietnam",
    "philippine airlines": "Philippines", "cebu pacific": "Philippines",
    "garuda indonesia": "Indonesia", "lion air": "Indonesia",
    "japan airlines": "Japan", "all nippon airways": "Japan",
    "korean air": "South Korea", "asiana airlines": "South Korea",
    "cathay pacific": "Hong Kong", "hong kong airlines": "Hong Kong",
    "china airlines": "Taiwan", "eva air": "Taiwan",
    "air china": "China", "china eastern": "China", "china southern": "China",
    "hainan airlines": "China", "juneyao air": "China",
    "pakistan international airlines": "Pakistan",
    "srilankan airlines": "Sri Lanka", "biman bangladesh": "Bangladesh",
    "air astana": "Kazakhstan",
    "aeromexico": "Mexico", "volaris": "Mexico", "viva aerobus": "Mexico",
    "latam airlines": "Chile", "gol linhas": "Brazil",
    "azul airlines": "Brazil", "avianca": "Colombia",
}

# Other proper nouns whose embedded geography is not the article's location.
NON_LOCATION_PHRASES = (
    "north america", "south america", "latin america", "north american",
    "west virginia university", "indiana university", "georgia tech",
    "washington post", "washington examiner", "washington times",
    "new york times", "usa today", "china lake", "new mexico state",
    "india pale ale", "georgia institute of technology",
    "virginia tech", "ohio state university", "texas instruments",
)


# --- Whole-phrase matching ----------------------------------------------------
#
# Every gazetteer lookup MUST respect word boundaries. An earlier build used
# plain substring matching and tagged five SAM.gov "HANGAR FIRE SUPPRESSION
# SYSTEM RENOVATION" notices as Nevada, because "Reno" is a substring of
# "RENOvation". \b is not usable here since some entries contain "." and "'"
# (e.g. "f.e. warren air force base", "coeur d'alene"), so explicit
# alphanumeric lookarounds are used instead.

def _phrase_regex(phrases) -> re.Pattern:
    """Compile one alternation regex over phrases, longest match first."""
    ordered = sorted({p.lower() for p in phrases if p}, key=len, reverse=True)
    return re.compile(
        r'(?<![a-z0-9])(' + "|".join(re.escape(p) for p in ordered) + r')(?![a-z0-9])'
    )


_GAZETTEER_RX: re.Pattern | None = None
_AIRLINE_RX: re.Pattern | None = None
_MASK_RX: re.Pattern | None = None


def gazetteer_hits(text_lower: str) -> set[str]:
    """Whole-phrase airport / base / city names present in the text."""
    global _GAZETTEER_RX
    if _GAZETTEER_RX is None:
        _GAZETTEER_RX = _phrase_regex(AIRPORT_LOCATIONS.keys())
    if not text_lower:
        return set()
    return {m.group(1) for m in _GAZETTEER_RX.finditer(text_lower)}


def airline_hits(text_lower: str) -> set[str]:
    global _AIRLINE_RX
    if _AIRLINE_RX is None:
        _AIRLINE_RX = _phrase_regex(AIRLINE_HOME.keys())
    if not text_lower:
        return set()
    return {m.group(1) for m in _AIRLINE_RX.finditer(text_lower)}


def mask_org_names(text: str) -> str:
    """
    Blank out airline and organisation names so geography inside them is not
    mistaken for the article's own location.
    """
    global _MASK_RX
    if not text:
        return ""
    if _MASK_RX is None:
        _MASK_RX = _phrase_regex(list(AIRLINE_HOME) + list(NON_LOCATION_PHRASES))
    return _MASK_RX.sub(" ", text)

# Distinctive cities. Only cities that map to exactly one country worldwide
# AND do not collide with a US city of the same name.
COUNTRY_CITIES = {
    "United Arab Emirates": ["Dubai", "Abu Dhabi", "Sharjah", "Ras Al Khaimah", "Ajman", "Fujairah"],
    "Saudi Arabia": ["Riyadh", "Jeddah", "Dammam", "Mecca", "Medina", "Tabuk", "NEOM", "Yanbu"],
    "Qatar": ["Doha", "Lusail"],
    "Bahrain": ["Manama"],
    "Kuwait": ["Kuwait City"],
    "Oman": ["Muscat", "Salalah", "Duqm"],
    "Israel": ["Tel Aviv", "Jerusalem", "Haifa"],
    "Egypt": ["Cairo", "Alexandria", "Giza", "Hurghada"],
    "Nigeria": ["Lagos", "Abuja", "Port Harcourt", "Ibadan", "Kano", "Enugu"],
    "Kenya": ["Nairobi", "Mombasa"],
    "South Africa": ["Johannesburg", "Cape Town", "Pretoria", "Durban", "Centurion", "Lanseria"],
    "Morocco": ["Casablanca", "Rabat", "Marrakech", "Marrakesh", "Tangier"],
    "Tunisia": ["Tunis"], "Algeria": ["Algiers"], "Libya": ["Tripoli", "Benghazi"],
    "Ethiopia": ["Addis Ababa"], "Tanzania": ["Dar es Salaam", "Arusha"],
    "Uganda": ["Kampala", "Entebbe"], "Senegal": ["Dakar"], "Ghana": ["Accra", "Kumasi"],
    "Rwanda": ["Kigali"], "Zimbabwe": ["Harare", "Bulawayo"], "Zambia": ["Lusaka"],
    "Botswana": ["Gaborone"], "Namibia": ["Windhoek"], "Mozambique": ["Maputo"],
    "Angola": ["Luanda"], "Ivory Coast": ["Abidjan", "Yamoussoukro"],
    "France": ["Toulouse", "Marseille", "Lyon", "Bordeaux", "Nantes", "Strasbourg", "Le Bourget"],
    "Germany": ["Berlin", "Munich", "Frankfurt", "Hamburg", "Cologne", "Stuttgart", "Düsseldorf", "Dusseldorf", "Leipzig", "Bremen", "Nuremberg"],
    "Spain": ["Madrid", "Barcelona", "Seville", "Sevilla", "Zaragoza", "Malaga", "Bilbao"],
    "Italy": ["Rome", "Milan", "Naples", "Turin", "Venice", "Bologna", "Genoa"],
    "Netherlands": ["Amsterdam", "Rotterdam", "Schiphol", "Eindhoven", "Utrecht", "The Hague"],
    "Belgium": ["Brussels", "Antwerp", "Liege", "Ghent", "Charleroi"],
    "Austria": ["Vienna", "Salzburg", "Graz", "Linz"],
    "Poland": ["Warsaw", "Krakow", "Kraków", "Wroclaw", "Gdansk", "Gdańsk", "Poznan", "Rzeszow", "Katowice", "Powidz"],
    "Czechia": ["Prague", "Brno", "Ostrava"],
    "Hungary": ["Budapest", "Debrecen"],
    "Romania": ["Bucharest", "Cluj", "Timisoara", "Constanta"],
    "Bulgaria": ["Sofia", "Plovdiv", "Varna", "Burgas"],
    "Greece": ["Athens", "Thessaloniki", "Heraklion"],
    "Portugal": ["Lisbon", "Porto", "Faro", "Beja"],
    "Sweden": ["Stockholm", "Gothenburg", "Malmo", "Linkoping"],
    "Norway": ["Oslo", "Bergen", "Stavanger", "Trondheim"],
    "Denmark": ["Copenhagen", "Aarhus", "Aalborg", "Billund"],
    "Finland": ["Helsinki", "Tampere", "Turku"],
    "Switzerland": ["Zurich", "Geneva", "Basel", "Bern", "Lausanne"],
    "Ireland": ["Dublin", "Shannon", "Galway"],
    "Cyprus": ["Nicosia", "Larnaca", "Limassol", "Paphos"],
    "Turkey": ["Istanbul", "Ankara", "Izmir", "Antalya", "Eskisehir"],
    "Ukraine": ["Kyiv", "Kiev", "Lviv", "Odesa", "Odessa", "Kharkiv"],
    "Croatia": ["Zagreb", "Split", "Dubrovnik"],
    "Serbia": ["Belgrade", "Novi Sad"],
    "Slovakia": ["Bratislava", "Kosice"],
    "Slovenia": ["Ljubljana", "Maribor"],
    "Lithuania": ["Vilnius", "Kaunas", "Siauliai"],
    "Latvia": ["Riga"], "Estonia": ["Tallinn", "Tartu"],
    "Iceland": ["Reykjavik", "Keflavik"],
    "Russia": ["Moscow", "Saint Petersburg", "Novosibirsk", "Kazan"],
    "United Kingdom": ["Farnborough", "Cambridgeshire", "Yorkshire", "Lancashire", "Hertfordshire", "Oxfordshire", "Gloucestershire", "Bedfordshire", "Warwickshire", "Cornwall", "Aberdeen", "Belfast", "Cardiff", "Swansea", "Coventry", "Sheffield", "Nottingham", "Leicester", "Bristol", "Brighton", "Southampton", "Luton", "Gatwick", "Heathrow", "Stansted"],
    "Australia": ["Sydney", "Melbourne", "Brisbane", "Perth", "Adelaide", "Canberra", "Darwin", "Hobart", "Townsville", "Toowoomba", "Geelong", "Newcastle NSW", "Wollongong"],
    "New Zealand": ["Auckland", "Wellington", "Christchurch", "Queenstown", "Dunedin", "Tauranga", "Palmerston North", "Ohakea", "Whenuapai"],
    "Japan": ["Tokyo", "Osaka", "Yokohama", "Nagoya", "Fukuoka", "Sapporo", "Kobe"],
    "Singapore": ["Changi", "Seletar"],
    "China": ["Beijing", "Shanghai", "Shenzhen", "Guangzhou", "Chengdu", "Tianjin", "Xiamen"],
    "Hong Kong": ["Hong Kong"],
    "India": ["Mumbai", "Bengaluru", "Bangalore", "Hyderabad", "Chennai", "Kolkata", "Pune", "Ahmedabad", "New Delhi", "Nagpur", "Kochi"],
    "Thailand": ["Bangkok", "Phuket", "Chiang Mai", "U-Tapao"],
    "Malaysia": ["Kuala Lumpur", "Subang", "Penang", "Johor Bahru"],
    "Indonesia": ["Jakarta", "Surabaya", "Bali", "Denpasar"],
    "Philippines": ["Manila", "Cebu", "Clark", "Davao"],
    "South Korea": ["Seoul", "Busan", "Incheon"],
    "Vietnam": ["Hanoi", "Ho Chi Minh", "Da Nang"],
    "Taiwan": ["Taipei", "Kaohsiung", "Taoyuan"],
    "Pakistan": ["Karachi", "Lahore", "Islamabad"],
    "Bangladesh": ["Dhaka", "Chittagong"],
    "Sri Lanka": ["Colombo"],
    "Kazakhstan": ["Almaty", "Astana"],
    "Uzbekistan": ["Tashkent"],
    "Brazil": ["Sao Paulo", "São Paulo", "Rio de Janeiro", "Brasilia", "Campinas"],
    "Mexico": ["Guadalajara", "Monterrey", "Tijuana", "Queretaro", "Cancun"],
    "Canada": ["Toronto", "Montreal", "Vancouver", "Calgary", "Edmonton", "Ottawa", "Winnipeg", "Halifax", "Saskatoon", "Regina", "Mississauga", "Mirabel", "Trenton Ontario"],
}


def strip_boilerplate(html: str) -> str:
    """
    Remove the parts of a page that pollute location detection: script, style,
    nav, header, footer, aside, form, noscript, and HTML comments.

    v1 skipped this step and fed raw tag-stripped HTML (mostly JavaScript) to
    the detector, which is the single biggest reason 55% of rows had no
    location and several had the wrong one.
    """
    if not html:
        return ""
    out = html
    out = re.sub(r'<!--.*?-->', ' ', out, flags=re.S)
    for tag in ("script", "style", "noscript", "nav", "header", "footer",
                "aside", "form", "svg", "iframe", "template"):
        out = re.sub(r'<' + tag + r'\b.*?</' + tag + r'\s*>', ' ', out,
                     flags=re.S | re.I)
    return out


def extract_article_text(html: str, limit: int = 6000) -> str:
    """
    Pull the readable article body: paragraph text from the de-boilerplated
    HTML, longest-first, so a cookie banner can't outrank the lede.
    """
    cleaned = strip_boilerplate(html)
    paras = re.findall(r'<p\b[^>]*>(.*?)</p>', cleaned, flags=re.S | re.I)
    texts = []
    for p in paras:
        t = re.sub(r'<[^>]+>', ' ', p)
        t = unescape_entities(re.sub(r'\s+', ' ', t)).strip()
        if len(t) >= 40:            # skip captions, bylines, share prompts
            texts.append(t)
    body = " ".join(texts)
    if len(body) < 200:
        # Fall back to all visible text if the page doesn't use <p> tags.
        body = re.sub(r'<[^>]+>', ' ', cleaned)
        body = unescape_entities(re.sub(r'\s+', ' ', body)).strip()
    return body[:limit]


def unescape_entities(text: str) -> str:
    """Decode HTML entities. v1 left '&amp;' visible in the Summary column."""
    import html as _html
    try:
        return _html.unescape(text)
    except Exception:
        return text


def _count(needle: str, haystack_lower: str) -> int:
    if not needle:
        return 0
    return len(re.findall(r'\b' + re.escape(needle.lower()) + r'\b', haystack_lower))


def _has(needle: str, haystack_lower: str) -> bool:
    return _count(needle, haystack_lower) > 0


# US-specific corroborating evidence, used to settle ambiguous names.
US_EVIDENCE = (
    "u.s.", "us air force", "united states", "air force base", "afb",
    "space force", "national guard", "pentagon", "faa", "usace",
    "army corps of engineers", "nfpa", "county", "city council",
    "state of", "governor", "senator", "congressman", "milcon",
    "department of defense", "naval air station", "marine corps",
    "$", "sam.gov", "gsa", "tsa", "ntsb",
)


def _us_evidence_count(blob: str) -> int:
    return sum(1 for e in US_EVIDENCE if e in blob)


def find_us_state(title_l: str, body_l: str) -> tuple[str | None, float]:
    """
    Find the most likely US state. Returns (state_name, score).

    Handles three patterns, strongest first:
      1. "City, TX" / "City, TX 75001"  -- very reliable in US news
      2. Full state name, title weighted 5x
      3. Bare two-letter code, only with other US evidence present
    """
    best, best_score = None, 0.0

    # 1. "Something, ST" — a comma followed by a state abbreviation
    for text, weight in ((title_l, 8.0), (body_l, 3.0)):
        for m in re.finditer(r',\s*([a-z]{2})\b', text):
            abbr = m.group(1).upper()
            if abbr in US_ABBR_TO_STATE:
                st = US_ABBR_TO_STATE[abbr]
                if weight > best_score:
                    best, best_score = st, weight

    # 2. Full state names
    for st in US_STATES:
        if st == "Georgia":
            continue        # handled by disambiguation, not here
        s = _count(st, title_l) * 6.0 + _count(st, body_l) * 1.5
        # "Washington" alone usually means DC or the newspaper, not the state
        if st == "Washington" and not _has("washington state", title_l + " " + body_l):
            s *= 0.4
        if s > best_score:
            best, best_score = st, s

    # 3. Georgia the US state, only with corroboration
    ga_hits = _count("georgia", title_l) * 6.0 + _count("georgia", body_l) * 1.5
    if ga_hits:
        blob = title_l + " " + body_l
        us_like = _us_evidence_count(blob) >= 1 or _has("atlanta", blob) or _has("robins", blob) or _has("savannah", blob)
        gia_like = _has("tbilisi", blob) or _has("caucasus", blob) or _has("georgian", blob)
        if us_like and not gia_like and ga_hits > best_score:
            best, best_score = "Georgia", ga_hits

    return best, best_score


def find_subdivision(names: dict | set, title_l: str, body_l: str) -> tuple[str | None, float]:
    """Generic best-scoring subdivision finder for Canada / Australia."""
    best, best_score = None, 0.0
    for name in names:
        s = _count(name, title_l) * 6.0 + _count(name, body_l) * 1.5
        if s > best_score:
            best, best_score = name, s
    return best, best_score


def detect_location(text: str, title: str = "", url: str = "") -> str:
    """
    Determine an article's location. Returns e.g. "United States - Texas",
    "Canada - Ontario", "Australia - Queensland", "United Arab Emirates",
    or "Unknown".

    Evidence is gathered in priority order and scored. The airport gazetteer
    and title hits dominate; the publisher's domain only breaks ties.
    """
    raw_title_l = (title or "").lower()
    raw_body_l = (text or "").lower()
    if not (raw_title_l + raw_body_l).strip():
        return UNKNOWN

    # Airline / org names are removed before any country, state or city
    # counting, so "Alaska Airlines" cannot vote for Alaska.
    title_l = mask_org_names(raw_title_l)
    body_l = mask_org_names(raw_body_l)
    blob = title_l + " " + body_l

    scores: dict[str, float] = {}
    subdivision: dict[str, str] = {}   # country -> state/province

    def bump(loc: str, pts: float) -> None:
        if pts <= 0 or not loc:
            return
        country = loc.split(" - ", 1)[0]
        if " - " in loc:
            sub = loc.split(" - ", 1)[1]
            # Remember the strongest subdivision seen for this country
            key = country + "||" + sub
            subdivision[key] = sub
            if scores.get(country + "##sub##" + sub, 0) < pts:
                scores[country + "##sub##" + sub] = pts
        scores[country] = scores.get(country, 0.0) + pts

    # --- 1. Airport / air base gazetteer (highest signal in this domain) ----
    #        Matched against the UNMASKED text: airport names are long and
    #        specific enough that masking would only lose information.
    title_gaz = gazetteer_hits(raw_title_l)
    body_gaz = gazetteer_hits(raw_body_l)
    for name in title_gaz:
        bump(AIRPORT_LOCATIONS[name], 14.0)
    for name in body_gaz - title_gaz:
        bump(AIRPORT_LOCATIONS[name], 5.0)

    # --- 1b. Airline home country — added back after masking. A carrier named
    #         in the HEADLINE clears the confidence floor on its own, since
    #         "Air Peace breaks ground on West Africa's largest MRO" is
    #         obviously Nigerian even with no place name in the text.
    title_air = airline_hits(raw_title_l)
    body_air = airline_hits(raw_body_l)
    for name in title_air:
        bump(AIRLINE_HOME[name], 5.0)
    for name in body_air - title_air:
        bump(AIRLINE_HOME[name], 1.0)

    # --- 2. Unambiguous country names --------------------------------------
    for country in ALL_COUNTRIES:
        if country in AMBIGUOUS_COUNTRY_NAMES:
            continue
        s = _count(country, title_l) * 10.0 + _count(country, body_l) * 2.0
        bump(country, s)

    # --- 3. Aliases and demonyms ------------------------------------------
    for alias, canonical in COUNTRY_ALIASES.items():
        s = _count(alias, title_l) * 7.0 + _count(alias, body_l) * 1.5
        bump(canonical, s)

    # --- 4. Ambiguous country names, only with corroboration ---------------
    #        e.g. "Georgia" needs Tbilisi/Caucasus; "China" must not be
    #        "China Lake"; "Turkey" must not be about food.
    for country in AMBIGUOUS_COUNTRY_NAMES:
        if country not in ALL_COUNTRIES:
            continue
        hits = _count(country, title_l) * 8.0 + _count(country, body_l) * 1.5
        if not hits:
            continue
        if country == "Georgia":
            if _has("tbilisi", blob) or _has("caucasus", blob) or _has("georgian", blob):
                bump("Georgia", hits)
            continue
        if country == "China" and _has("china lake", blob):
            continue
        if country == "Turkey" and not (_has("turkish", blob) or _has("ankara", blob) or _has("istanbul", blob)):
            hits *= 0.3
        if country == "Mexico" and _has("new mexico", blob):
            continue
        if country == "India" and _has("indiana", blob):
            hits *= 0.3
        if country == "Lebanon" and _has("beirut", blob) is False and _us_evidence_count(blob) >= 2:
            continue
        bump(country, hits)

    # --- 5. Distinctive cities --------------------------------------------
    for country, cities in COUNTRY_CITIES.items():
        for city in cities:
            s = _count(city, title_l) * 8.0 + _count(city, body_l) * 1.5
            bump(country, s)

    # --- 6. US states -----------------------------------------------------
    us_state, us_score = find_us_state(title_l, body_l)
    if us_score > 0:
        bump("United States - " + us_state if us_state else "United States", us_score)

    # --- 7. Canadian provinces --------------------------------------------
    ca_prov, ca_score = find_subdivision(set(CA_PROVINCES), title_l, body_l)
    if ca_score > 0 and ca_prov:
        # Normalise "Québec" -> "Quebec" for consistent tab grouping
        canon_prov = "Quebec" if ca_prov == "Québec" else ca_prov
        bump("Canada - " + canon_prov, ca_score)
    for m in re.finditer(r',\s*([A-Za-z]{2})\b', title or ""):
        abbr = m.group(1).upper()
        if abbr in CA_ABBR_TO_PROVINCE and abbr not in US_ABBR_TO_STATE:
            bump("Canada - " + CA_ABBR_TO_PROVINCE[abbr], 8.0)

    # --- 8. Australian states ---------------------------------------------
    au_state, au_score = find_subdivision(set(AU_STATES), title_l, body_l)
    if au_score > 0:
        # "Victoria" and "Western Australia" collide with other things;
        # require an Australia signal for the weaker ones.
        au_ok = (au_state not in ("Victoria",)) or _has("australia", blob)
        if au_ok:
            bump("Australia - " + au_state, au_score)

    # --- 9. New Zealand regions -------------------------------------------
    nz_region, nz_score = find_subdivision(NZ_REGIONS, title_l, body_l)
    if nz_score > 0 and (_has("new zealand", blob) or _has("nz", blob) or nz_region in ("Auckland", "Wellington", "Canterbury", "Otago", "Waikato")):
        bump("New Zealand", nz_score)

    # --- 10. Mexican states -----------------------------------------------
    mx_state, mx_score = find_subdivision(MX_STATES, title_l, body_l)
    if mx_score > 0 and _has("mexico", blob):
        bump("Mexico", mx_score)

    # --- 11. Publisher evidence: domain map, then ccTLD -------------------
    host = registrable_domain(url)
    if host:
        for dom, country in DOMAIN_COUNTRY.items():
            if host == dom or host.endswith("." + dom):
                if country in ALL_COUNTRIES:
                    bump(country, 3.0)
                break
        for tld, country in sorted(TLD_COUNTRY.items(), key=lambda kv: -len(kv[0])):
            if host.endswith(tld):
                bump(country, 2.5)
                break
        if host.endswith(".gov") or host.endswith(".mil") or host.endswith(".us"):
            bump("United States", 3.0)

    # --- 12. Broad regional terms (routes to a tab even without a country) --
    for term, region_name in REGION_TERMS.items():
        s_ = _count(term, title_l) * 6.0 + _count(term, body_l) * 1.5
        if s_:
            scores[region_name] = scores.get(region_name, 0.0) + s_

    # --- Decide ------------------------------------------------------------
    country_scores = {k: v for k, v in scores.items() if "##sub##" not in k}

    best_country = max(country_scores, key=country_scores.get) if country_scores else ""
    best_score = country_scores.get(best_country, 0.0)

    # Confidence floor. Below this, publisher-TLD noise is all we have.
    if best_score < 4.0:
        # Last resort: a US federal agency in the text means the project is
        # American even when no state or city is named. Better to file it under
        # USA than to strand a Coast Guard hangar award in Needs Review.
        if any(m in blob for m in US_FEDERAL_MARKERS) and _us_evidence_count(blob) >= 1:
            return "United States"
        return UNKNOWN

    # Attach the strongest subdivision belonging to the winning country.
    subs = {k.split("##sub##", 1)[1]: v
            for k, v in scores.items()
            if k.startswith(best_country + "##sub##")}
    if subs:
        best_sub = max(subs, key=subs.get)
        return best_country + " - " + best_sub
    return best_country


# ==============================================================================
# REGION / TAB ROUTING
# ==============================================================================

REGION_ORDER = ["USA", "CANADA", "ANZ", "EMEA", "REVIEW"]
REGION_TAB_NAMES = {
    "USA": "USA",
    "CANADA": "Canada",
    "ANZ": "Australia-NZ",
    "EMEA": "EMEA",
    "REVIEW": "Needs Review",
}


def classify_region(location: str) -> str:
    """
    Route a location string to a tab. Unlike v1, an unknown location goes to
    REVIEW rather than silently defaulting to the North America tab.
    """
    if not location or location == UNKNOWN:
        return "REVIEW"
    base = location.split(" - ", 1)[0].strip()
    region = COUNTRY_REGION.get(base)
    if region:
        return region
    low = base.lower()
    if low in ("europe", "european union", "eu", "emea", "middle east",
               "africa", "asia", "asia pacific", "apac"):
        return "EMEA"
    if low in ("australia", "new zealand", "oceania"):
        return "ANZ"
    if low in ("united states", "usa", "u.s."):
        return "USA"
    if low == "canada":
        return "CANADA"
    return "REVIEW"


def split_by_region(rows: list[dict]) -> dict[str, list[dict]]:
    buckets: dict[str, list[dict]] = {r: [] for r in REGION_ORDER}
    for row in rows:
        buckets[classify_region(row.get("Location", ""))].append(row)
    return buckets


# ==============================================================================
# DATES
# ==============================================================================

def today() -> datetime.date:
    return datetime.date.today()


def date_range(days: int = LOOKBACK_DAYS) -> tuple[str, str]:
    end = today()
    start = end - datetime.timedelta(days=days)
    return start.isoformat(), end.isoformat()


def week_label() -> str:
    t = today()
    return (t - datetime.timedelta(days=t.weekday())).isoformat()


def parse_any_date(raw: str) -> str:
    """
    Normalise a date string to YYYY-MM-DD, or "" if it can't be parsed.

    Unlike v1 this does NOT fall back to scraping any ISO-looking string out
    of the page HTML. That fallback is what produced a row dated 2026-08-13,
    ten days in the future, in the 8/3 report.
    """
    if not raw:
        return ""
    raw = str(raw).strip()

    m = re.match(r'(\d{4})-(\d{2})-(\d{2})', raw)
    if m:
        return m.group(0)

    # Relative phrasing from Google ("3 days ago", "2 weeks ago")
    low = raw.lower()
    if "ago" in low:
        n_match = re.search(r'(\d+)', low)
        n = int(n_match.group(1)) if n_match else 1
        if "hour" in low or "minute" in low or "second" in low:
            return today().isoformat()
        if "day" in low:
            return (today() - datetime.timedelta(days=n)).isoformat()
        if "week" in low:
            return (today() - datetime.timedelta(weeks=n)).isoformat()
        if "month" in low:
            return (today() - datetime.timedelta(days=30 * n)).isoformat()
        if "year" in low:
            return (today() - datetime.timedelta(days=365 * n)).isoformat()

    cleaned = re.sub(r'\s+', ' ', raw.replace(",", ", ")).strip(" ,")
    for fmt in ("%B %d, %Y", "%b %d, %Y", "%d %B %Y", "%d %b %Y",
                "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%d.%m.%Y", "%Y%m%d"):
        try:
            return datetime.datetime.strptime(cleaned, fmt).date().isoformat()
        except ValueError:
            continue
    return ""


def date_is_sane(date_str: str, max_age_days: int) -> bool:
    """A date must be parseable, not in the future, and inside the window."""
    if not date_str:
        return False
    try:
        d = datetime.date.fromisoformat(date_str)
    except ValueError:
        return False
    t = today()
    if d > t + datetime.timedelta(days=1):    # tolerate timezone edge
        return False
    return (t - d).days <= max_age_days


# ==============================================================================
# ARTICLE PAGE ENRICHMENT
# ==============================================================================

META_DATE_PROPS = [
    "article:published_time", "og:published_time", "datepublished",
    "publish_date", "pubdate", "date", "dc.date", "dc.date.issued",
    "sailthru.date", "parsely-pub-date", "article:modified_time",
]
JSONLD_DATE_KEYS = ["datePublished", "dateCreated", "dateModified"]


def parse_meta(html: str) -> dict:
    """
    Extract <meta> content, <title>, and JSON-LD blocks with regex.

    v1 used html.parser.HTMLParser. That works, but a regex pass is more
    tolerant of the malformed markup that news sites routinely ship, and it
    lets us operate on the de-boilerplated HTML.
    """
    meta: dict[str, str] = {}
    for m in re.finditer(r'<meta\b([^>]*)>', html, flags=re.I):
        attrs = m.group(1)
        key = re.search(r'(?:property|name|itemprop)\s*=\s*["\']([^"\']+)["\']', attrs, flags=re.I)
        val = re.search(r'content\s*=\s*["\']([^"\']*)["\']', attrs, flags=re.I)
        if key and val and val.group(1).strip():
            meta[key.group(1).strip().lower()] = val.group(1).strip()

    title_m = re.search(r'<title\b[^>]*>(.*?)</title>', html, flags=re.S | re.I)
    title = unescape_entities(re.sub(r'\s+', ' ', title_m.group(1))).strip() if title_m else ""

    scripts = re.findall(
        r'<script\b[^>]*type\s*=\s*["\']application/ld\+json["\'][^>]*>(.*?)</script>',
        html, flags=re.S | re.I)

    return {"meta": meta, "title": title, "scripts": scripts}


def date_from_jsonld(scripts: list[str]) -> str:
    for raw in scripts:
        try:
            data = json.loads(raw.strip())
        except Exception:
            continue
        stack = [data]
        while stack:
            node = stack.pop()
            if isinstance(node, dict):
                for key in JSONLD_DATE_KEYS:
                    if node.get(key):
                        got = parse_any_date(node[key])
                        if got:
                            return got
                stack.extend(node.values())
            elif isinstance(node, list):
                stack.extend(node)
    return ""


def fetch_article_meta(url: str) -> dict:
    """
    Fetch an article and extract an accurate publish date and location.
    Returns {"publish_date": str, "location": str, "fetched": bool}.
    """
    result = {"publish_date": "", "location": "", "fetched": False}
    try:
        r = requests.get(url, headers=HEADERS_UA, timeout=15, allow_redirects=True)
        if r.status_code != 200 or not r.text:
            return result
        html = r.text[:400_000]
    except Exception as exc:
        log.debug("Could not fetch %s: %s", url, exc)
        return result

    result["fetched"] = True
    parsed = parse_meta(html)
    meta, page_title, scripts = parsed["meta"], parsed["title"], parsed["scripts"]

    # Date: meta tags, then JSON-LD. No blind regex scrape of the whole page.
    for prop in META_DATE_PROPS:
        if meta.get(prop):
            got = parse_any_date(meta[prop])
            if got:
                result["publish_date"] = got
                break
    if not result["publish_date"]:
        result["publish_date"] = date_from_jsonld(scripts)
    if not result["publish_date"]:
        # Narrow, targeted fallback: a <time datetime="..."> element only.
        m = re.search(r'<time\b[^>]*datetime\s*=\s*["\']([^"\']+)["\']', html, flags=re.I)
        if m:
            result["publish_date"] = parse_any_date(m.group(1))

    # Location: real article text, plus the strongest title we can find.
    best_title = (meta.get("og:title") or meta.get("twitter:title")
                  or page_title or "")
    body = extract_article_text(html)
    desc = unescape_entities(meta.get("og:description") or meta.get("description") or "")
    # Some sites put the datelined city in these fields — very useful.
    geo_hints = " ".join(filter(None, [
        meta.get("geo.region", ""), meta.get("geo.placename", ""),
        meta.get("article:section", ""), meta.get("og:locale", ""),
    ]))
    result["location"] = detect_location(
        " ".join([desc, geo_hints, body]), title=best_title, url=url)
    return result


def enrich_with_page_data(rows: list[dict]) -> list[dict]:
    """
    Visit each article and replace the search-snippet date/location with
    values read from the page itself.

    v1 gave every un-dated news row a fake "~today" date. That marked 156 of
    341 rows (46%) in the 8/3 report as if they were published that week.
    v2 leaves the search engine's own date in place instead, and rows that
    still have no credible date are dropped by the sanity filter.
    """
    total = len(rows)
    for i, row in enumerate(rows, start=1):
        url = row.get("Source URL", "")
        if not url or row.get("Source") != "news":
            continue
        log.info("Enriching %d/%d: %s", i, total, url[:110])
        meta = fetch_article_meta(url)
        if meta["publish_date"]:
            row["Date Published"] = meta["publish_date"]
        # Only overwrite the location if the page gave us a real answer.
        if meta["location"] and meta["location"] != UNKNOWN:
            row["Location"] = meta["location"]
        elif not row.get("Location"):
            row["Location"] = UNKNOWN
        time.sleep(0.4)
    return rows


# ==============================================================================
# SERPAPI
# ==============================================================================

SERPAPI_SEARCH_URL = "https://serpapi.com/search"
SERPAPI_ACCOUNT_URL = "https://serpapi.com/account.json"

_SERPAPI_ERRORS = 0
_SERPAPI_CALLS = 0


def check_serpapi_quota() -> dict:
    """Free endpoint — does not consume search budget."""
    try:
        r = requests.get(SERPAPI_ACCOUNT_URL,
                         params={"api_key": env("SERPAPI_KEY")}, timeout=15)
        r.raise_for_status()
        return r.json()
    except Exception as exc:
        log.warning("Could not check SerpAPI quota: %s", exc)
        return {}


def log_quota_status(label: str, quota: dict) -> None:
    if not quota:
        log.info("[%s] SerpAPI quota: <unavailable>", label)
        return
    log.info(
        "[%s] SerpAPI — plan=%s | %s/%s used this month | %s left | %s/%s last hour",
        label, quota.get("plan_name", "?"), quota.get("this_month_usage", "?"),
        quota.get("searches_per_month", "?"), quota.get("total_searches_left", "?"),
        quota.get("last_hour_searches", "?"),
        quota.get("account_rate_limit_per_hour", "?"),
    )


def serpapi_news_search(query: str) -> list[dict]:
    """
    Search the Google NEWS tab, restricted to the past week.

    THE KEY v2 CHANGE. v1 called engine=google with no tbm and no tbs, i.e.
    an ordinary web search over all of history. That is why the 8/3 report
    contained vendor brochures, YouTube videos, Reddit threads, Wikipedia
    articles and pages from 2025 — they are legitimately the best *web*
    results for "aircraft hangar construction", just not news.
    """
    global _SERPAPI_ERRORS, _SERPAPI_CALLS
    params = {
        "engine": "google",
        "tbm": "nws",           # News tab
        "tbs": "qdr:w",         # past week only
        "q": query,
        "gl": "us",
        "hl": "en",
        "num": 100,             # SerpAPI bills per search, not per result
        "api_key": env("SERPAPI_KEY"),
    }
    try:
        _SERPAPI_CALLS += 1
        r = requests.get(SERPAPI_SEARCH_URL, params=params, timeout=30)

        if r.status_code == 429:
            _SERPAPI_ERRORS += 1
            log.error(
                "SerpAPI RATE LIMITED for '%s' (HTTP 429). Free tier allows "
                "50 searches/hour. Either a run was re-triggered too soon, or "
                "the monthly cap is exhausted. See https://serpapi.com/dashboard",
                query)
            return []

        r.raise_for_status()
        data = r.json()

        if "error" in data:
            msg = data["error"]
            if any(t in msg.lower() for t in ("run out", "exceeded", "limit", "quota")):
                _SERPAPI_ERRORS += 1
                log.error("SerpAPI QUOTA ERROR for '%s': %s", query, msg)
            else:
                log.warning("SerpAPI error for '%s': %s", query, msg)
            return []

        results = data.get("news_results") or data.get("organic_results") or []
        log.info("SerpAPI news '%s' -> %d results", query, len(results))
        return results
    except Exception as exc:
        log.warning("SerpAPI request failed for '%s': %s", query, exc)
        return []


def parse_news_result(item: dict) -> dict | None:
    title = unescape_entities((item.get("title") or "").strip())
    link = (item.get("link") or "").strip()
    snippet = unescape_entities((item.get("snippet") or item.get("description") or "").strip())
    if not title or not link:
        return None

    raw_date = item.get("date") or item.get("published_at") or item.get("iso_date") or ""
    published = parse_any_date(raw_date)

    # The News tab gives us the publisher name — useful location evidence.
    source_name = ""
    src = item.get("source")
    if isinstance(src, dict):
        source_name = src.get("name", "") or ""
    elif isinstance(src, str):
        source_name = src

    return {
        "Project Title": title,
        "Source URL": link,
        "Summary": snippet[:400],
        "Date Published": published,
        "Location": detect_location(snippet + " " + source_name, title=title, url=link),
        "Source": "news",
    }


# ==============================================================================
# PROCUREMENT SOURCES
# ==============================================================================
#
# None of these need an API key. If you previously hunted for a developer
# registration page on AusTender or CanadaBuys and came up empty, that is
# because there is no key to get — the data is open.

HANGAR_KEYWORDS = (
    "hangar", "aircraft", "airfield", "aviation", "airport", "aerodrome",
    "fire suppression", "fire protection", "foam system", "afff",
    "suppression system", "deluge",
)

# ---- SAM.gov (United States) ------------------------------------------------

SAM_API_URL = "https://api.sam.gov/opportunities/v2/search"
SAM_TITLE_KEYWORDS = [
    "hangar", "aircraft hangar", "hangar fire suppression",
    "hangar fire protection", "hangar construction", "hangar renovation",
    "aircraft maintenance facility",
]


def samgov_search(start_date: str, end_date: str) -> list[dict]:
    def mmddyyyy(d: str) -> str:
        y, m, dd = d.split("-")
        return f"{m}/{dd}/{y}"

    results, seen = [], set()
    for keyword in SAM_TITLE_KEYWORDS:
        params = {
            "api_key": env("SAM_API_KEY", "DEMO_KEY"),
            "postedFrom": mmddyyyy(start_date),
            "postedTo": mmddyyyy(end_date),
            "title": keyword,
            "limit": 25,
            "offset": 0,
        }
        try:
            r = requests.get(SAM_API_URL, params=params, timeout=40)
            if r.status_code in (401, 403):
                log.error(
                    "SAM.gov returned %d — the SAM_API_KEY has most likely "
                    "expired (they rotate every 90 days). Renew it at "
                    "sam.gov/workspace/profile/account-details and update the "
                    "GitHub secret.", r.status_code)
                return results
            r.raise_for_status()
            opps = r.json().get("opportunitiesData", [])
            log.info("SAM.gov '%s' -> %d results", keyword, len(opps))
            for opp in opps:
                nid = opp.get("noticeId", "")
                if nid in seen:
                    continue
                seen.add(nid)
                row = parse_sam_result(opp)
                if row:
                    results.append(row)
        except Exception as exc:
            log.warning("SAM.gov error for '%s': %s", keyword, exc)
        time.sleep(0.5)
    return results


def parse_sam_result(opp: dict) -> dict | None:
    title = (opp.get("title") or "").strip()
    desc = (opp.get("description") or "").strip()
    if not any(kw in (title + " " + desc).lower() for kw in HANGAR_KEYWORDS):
        return None

    place = opp.get("placeOfPerformance") or {}
    state = ((place.get("state") or {}).get("name") or "").strip()
    city = ((place.get("city") or {}).get("name") or "").strip()
    country = ((place.get("country") or {}).get("name") or "United States").strip()

    if country in ("United States", "USA", "UNITED STATES", ""):
        location = "United States - " + state if state else "United States"
    else:
        # Overseas MILCON — let the detector place it in the host country.
        detected = detect_location(f"{city} {state} {country} {desc[:800]}", title=title)
        location = detected if detected != UNKNOWN else country

    return {
        "Project Title": title,
        "Source URL": "https://sam.gov/opp/" + opp.get("noticeId", "") + "/view",
        "Summary": re.sub(r'\s+', ' ', desc)[:400],
        "Date Published": (opp.get("postedDate") or "")[:10],
        "Location": location,
        "Source": "sam_gov",
    }


# ---- CanadaBuys (Canada) ----------------------------------------------------
#
# Open data CSV, no key. The v1 code pointed at the right URL, but the run
# produced ZERO Canadian rows. Two reasons, both fixed here:
#   (a) timeout=30 on a large CSV meant the whole request raised and was
#       swallowed by a bare `except`. This version streams with a 120s budget.
#   (b) the description column name was guessed wrong, so the keyword filter
#       only ever saw the title. This version discovers columns by pattern
#       instead of hard-coding names, so a schema change can't silently
#       zero out the source again.

CANADABUYS_CSV_URL = "https://canadabuys.canada.ca/opendata/pub/newTenderNotice-nouvelAvisAppelOffres.csv"


def _pick_column(fieldnames: list[str], *patterns: str) -> str:
    """Find the first column whose name matches any pattern (case-insensitive)."""
    for pat in patterns:
        rx = re.compile(pat, re.I)
        for name in fieldnames:
            if name and rx.search(name):
                return name
    return ""


def canadabuys_search(start_date: str) -> list[dict]:
    results: list[dict] = []
    try:
        r = requests.get(CANADABUYS_CSV_URL, headers=HEADERS_UA,
                         timeout=(15, 120), stream=True)
        r.raise_for_status()
        r.encoding = r.encoding or "utf-8"
        text = r.text
    except Exception as exc:
        log.warning("CanadaBuys download failed: %s", exc)
        return results

    try:
        reader = csv.DictReader(io.StringIO(text))
        cols = reader.fieldnames or []
        c_title = _pick_column(cols, r'^title.*eng', r'title.*titre.*eng', r'^title$')
        c_desc = _pick_column(cols, r'description.*eng', r'^description', r'tenderDescription')
        c_date = _pick_column(cols, r'publicationDate', r'publication.*date', r'datePublication', r'^date')
        c_ref = _pick_column(cols, r'referenceNumber', r'reference.*numero', r'^solicitationNumber')
        c_url = _pick_column(cols, r'noticeURL.*eng', r'notice.*url', r'^url')
        c_region = _pick_column(cols, r'regionsOfDelivery.*eng', r'regionsOfDelivery', r'deliveryRegion')
        log.info("CanadaBuys columns -> title=%r desc=%r date=%r ref=%r url=%r region=%r",
                 c_title, c_desc, c_date, c_ref, c_url, c_region)

        scanned = 0
        for rec in reader:
            scanned += 1
            title = (rec.get(c_title) or "").strip()
            if not title:
                continue
            desc = (rec.get(c_desc) or "").strip()
            if not any(kw in (title + " " + desc).lower() for kw in HANGAR_KEYWORDS):
                continue

            pub = parse_any_date((rec.get(c_date) or "")[:19])
            if pub and pub < start_date:
                continue

            ref = (rec.get(c_ref) or "").strip()
            url = (rec.get(c_url) or "").strip()
            if not url:
                url = ("https://canadabuys.canada.ca/en/tender-opportunities/tender-notice/" + ref
                       if ref else "https://canadabuys.canada.ca/en/tender-opportunities")

            region = (rec.get(c_region) or "").strip()
            location = "Canada"
            if region:
                detected = detect_location(region + " Canada", title=region)
                if detected.startswith("Canada"):
                    location = detected

            results.append({
                "Project Title": title,
                "Source URL": url,
                "Summary": re.sub(r'\s+', ' ', desc)[:400],
                "Date Published": pub,
                "Location": location,
                "Source": "canadabuys",
            })
        log.info("CanadaBuys -> %d relevant of %d notices scanned", len(results), scanned)
    except Exception as exc:
        log.warning("CanadaBuys parse error: %s", exc)
    return results


# ---- AusTender (Australia) --------------------------------------------------
#
# v1 called https://api.tenders.gov.au/atm/v1/releases with a Bearer token.
# That endpoint does not exist and no token is issued, which is why
# registration kept failing. The real, unauthenticated OCDS endpoint is:
#   https://api.tenders.gov.au/ocds/findByDates/contractPublished/{from}/{to}
# Dates are ISO 8601. Verified live 2026-08-12: HTTP 200, 50 releases.

AUSTENDER_OCDS_BASE = "https://api.tenders.gov.au/ocds/findByDates"


def austender_search(start_date: str, end_date: str) -> list[dict]:
    results: list[dict] = []
    frm = start_date + "T00:00:00Z"
    to = end_date + "T23:59:59Z"
    for kind in ("contractPublished", "contractLastModified"):
        url = f"{AUSTENDER_OCDS_BASE}/{kind}/{frm}/{to}"
        try:
            r = requests.get(url, headers={"Accept": "application/json", **HEADERS_UA}, timeout=60)
            r.raise_for_status()
            releases = r.json().get("releases", [])
            log.info("AusTender %s -> %d releases", kind, len(releases))
        except Exception as exc:
            log.warning("AusTender %s error: %s", kind, exc)
            continue

        for rel in releases:
            for contract in (rel.get("contracts") or [{}]):
                title = (contract.get("title") or "").strip()
                desc = (contract.get("description") or "").strip()
                tender = rel.get("tender") or {}
                if not desc:
                    desc = (tender.get("description") or "").strip()
                blob = (title + " " + desc).lower()
                if not any(kw in blob for kw in HANGAR_KEYWORDS):
                    continue
                # AusTender puts a reference number in `title`; the readable
                # text lives in `description`. Prefer whichever reads better.
                display = desc if (not title or title.isdigit() or len(title) < 12) else title
                ocid = rel.get("ocid") or rel.get("id") or ""
                cn = re.sub(r'[^A-Za-z0-9]', '', ocid.split("-")[-1]) if ocid else ""
                link = ("https://www.tenders.gov.au/Cn/Show/" + cn if cn
                        else "https://www.tenders.gov.au/cn/search")

                buyer_region = ""
                for party in (rel.get("parties") or []):
                    addr = (party.get("address") or {})
                    buyer_region = " ".join(filter(None, [
                        addr.get("region", ""), addr.get("locality", "")]))
                    if buyer_region:
                        break

                detected = detect_location(buyer_region + " Australia " + desc[:600],
                                           title=display)
                location = detected if detected.startswith("Australia") else "Australia"

                results.append({
                    "Project Title": display[:220],
                    "Source URL": link,
                    "Summary": re.sub(r'\s+', ' ', desc)[:400],
                    "Date Published": parse_any_date(rel.get("date", "")),
                    "Location": location,
                    "Source": "austender",
                })
        time.sleep(0.5)
    log.info("AusTender total -> %d relevant results", len(results))
    return results


# ---- New Zealand GETS ------------------------------------------------------
#
# You had not tried NZ yet. There is no API, but GETS publishes an RSS feed of
# all current tenders and it needs no key:
#   https://www.gets.govt.nz/ExternalRSSFeed.htm
# MBIE's bulk CSVs are the other option but they only refresh quarterly, which
# is far too stale for a weekly report.

NZ_GETS_RSS = "https://www.gets.govt.nz/ExternalRSSFeed.htm"


def nz_gets_search() -> list[dict]:
    results: list[dict] = []
    try:
        r = requests.get(NZ_GETS_RSS, headers=HEADERS_UA, timeout=45)
        r.raise_for_status()
        xml = r.text
    except Exception as exc:
        log.warning("NZ GETS error: %s", exc)
        return results

    items = re.findall(r'<item\b.*?</item>', xml, flags=re.S | re.I)
    log.info("NZ GETS feed -> %d items", len(items))

    def tag(block: str, name: str) -> str:
        m = re.search(rf'<{name}\b[^>]*>(.*?)</{name}>', block, flags=re.S | re.I)
        if not m:
            return ""
        val = m.group(1)
        val = re.sub(r'<!\[CDATA\[(.*?)\]\]>', r'\1', val, flags=re.S)
        return unescape_entities(re.sub(r'\s+', ' ', re.sub(r'<[^>]+>', ' ', val))).strip()

    for block in items:
        title = tag(block, "title")
        desc = tag(block, "description")
        if not title:
            continue
        if not any(kw in (title + " " + desc).lower() for kw in HANGAR_KEYWORDS):
            continue
        link = tag(block, "link") or "https://www.gets.govt.nz/"
        pub = parse_any_date(tag(block, "pubDate") or tag(block, "dc:date"))
        results.append({
            "Project Title": title[:220],
            "Source URL": link,
            "Summary": desc[:400],
            "Date Published": pub,
            "Location": "New Zealand",
            "Source": "nz_gets",
        })
    log.info("NZ GETS -> %d relevant results", len(results))
    return results


# ---- TED Europa ------------------------------------------------------------
#
# Returned 0 rows in the 8/3 run. The v3 search API expects `q` (an expert
# query string) rather than `query`, and the field list must use the API's own
# field identifiers. Both are corrected here, and a v2-style fallback URL is
# tried if v3 rejects the payload.

TED_API_URL = "https://api.ted.europa.eu/v3/notices/search"
TED_QUERIES = [
    'hangar', 'aircraft hangar', '"fire suppression" AND hangar',
    'MRO aircraft maintenance facility', 'airfield construction',
]


def ted_europa_search(start_date: str, end_date: str) -> list[dict]:
    results, seen = [], set()
    frm = start_date.replace("-", "")
    to = end_date.replace("-", "")

    for term in TED_QUERIES:
        payload = {
            "q": f'({term}) AND publication-date>={frm} AND publication-date<={to}',
            "fields": ["publication-number", "notice-title", "publication-date",
                       "buyer-country", "place-of-performance"],
            "page": 1,
            "limit": 50,
            "scope": "ALL",
        }
        try:
            r = requests.post(TED_API_URL, json=payload,
                              headers={"Content-Type": "application/json"}, timeout=45)
            if r.status_code >= 400:
                log.warning("TED '%s' -> HTTP %d %s", term, r.status_code, r.text[:200])
                continue
            notices = r.json().get("notices", [])
            log.info("TED Europa '%s' -> %d notices", term, len(notices))
        except Exception as exc:
            log.warning("TED Europa error for '%s': %s", term, exc)
            continue

        for n in notices:
            pub_no = str(n.get("publication-number") or n.get("ND") or "")
            if not pub_no or pub_no in seen:
                continue
            seen.add(pub_no)

            title = _ted_text(n.get("notice-title"))
            if not title:
                continue
            country = _ted_text(n.get("buyer-country"))
            place = _ted_text(n.get("place-of-performance"))
            pub = parse_any_date(_ted_text(n.get("publication-date")))

            detected = detect_location(f"{place} {country}", title=title)
            location = detected if detected != UNKNOWN else (country or "Europe")

            results.append({
                "Project Title": title[:220],
                "Source URL": f"https://ted.europa.eu/en/notice/-/detail/{pub_no}",
                "Summary": re.sub(r'\s+', ' ', f"{place} {country}").strip()[:400],
                "Date Published": pub,
                "Location": location,
                "Source": "ted_europa",
            })
        time.sleep(0.6)

    log.info("TED Europa total -> %d results", len(results))
    return results


def _ted_text(value) -> str:
    """TED returns multilingual dicts and lists. Prefer English."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, list):
        return " ".join(_ted_text(v) for v in value).strip()
    if isinstance(value, dict):
        for key in ("eng", "en", "ENG", "EN"):
            if value.get(key):
                return _ted_text(value[key])
        return _ted_text(next(iter(value.values()), ""))
    return str(value)


# ==============================================================================
# DEDUPLICATION
# ==============================================================================

def title_fingerprint(title: str) -> str:
    """
    Normalised title for near-duplicate detection. Strips punctuation,
    stop-words and the trailing publisher name that news sites append
    ("... | Aviation Week"), so the same story syndicated to three outlets
    collapses to one row.
    """
    t = (title or "").lower()
    t = re.split(r'\s+[\|–—-]\s+', t)[0]        # drop " | Publisher"
    t = re.sub(r'[^a-z0-9 ]', ' ', t)
    words = [w for w in t.split() if len(w) > 2 and w not in {
        "the", "and", "for", "with", "new", "its", "has", "will", "from",
        "that", "this", "are", "was", "were", "been"}]
    return " ".join(sorted(set(words))[:12])


def deduplicate(rows: list[dict]) -> list[dict]:
    seen_urls: set[str] = set()
    seen_titles: set[str] = set()
    unique: list[dict] = []
    for row in rows:
        u = normalise_url(row.get("Source URL", ""))
        t = title_fingerprint(row.get("Project Title", ""))
        if (u and u in seen_urls) or (t and t in seen_titles):
            continue
        if u:
            seen_urls.add(u)
        if t:
            seen_titles.add(t)
        unique.append(row)
    return unique


# ==============================================================================
# CUMULATIVE HISTORY
# ==============================================================================
#
# The report is cumulative: each week's finds are added ABOVE the prior weeks'
# rows so sales never has to open an old email to find something. State lives
# in history.csv, committed back to the repo by the GitHub Actions workflow.
#
# `First Seen` is stored in the CSV only. It drives the "newest batch on top"
# ordering but is deliberately NOT written to the spreadsheet, because the
# column set is staying as-is.

HISTORY_FIELDS = ["Project Title", "Source URL", "Summary", "Date Published",
                  "Location", "Source", "First Seen", "Score"]


def load_history(path: str = None) -> list[dict]:
    path = path or HISTORY_FILE
    if not os.path.exists(path):
        log.info("No history file at %s — starting fresh.", path)
        return []
    rows = []
    try:
        with open(path, "r", encoding="utf-8", newline="") as fh:
            for rec in csv.DictReader(fh):
                rows.append({k: (rec.get(k) or "") for k in HISTORY_FIELDS})
        log.info("Loaded %d rows from %s", len(rows), path)
    except Exception as exc:
        log.warning("Could not read history %s: %s", path, exc)
    return rows


def save_history(rows: list[dict], path: str = None) -> None:
    path = path or HISTORY_FILE
    try:
        with open(path, "w", encoding="utf-8", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=HISTORY_FIELDS)
            writer.writeheader()
            for row in rows:
                writer.writerow({k: row.get(k, "") for k in HISTORY_FIELDS})
        log.info("Wrote %d rows to %s", len(rows), path)
    except Exception as exc:
        log.error("Could not write history %s: %s", path, exc)


def merge_history(history: list[dict], new_rows: list[dict],
                  run_date: str) -> tuple[list[dict], list[dict]]:
    """
    Add this run's rows to the accumulated history.

    Returns (merged, genuinely_new). A row already present in history — by
    normalised URL or by title fingerprint — is skipped, so an article that
    keeps resurfacing in search results does not reappear in the report.
    """
    known_urls = {normalise_url(r.get("Source URL", "")) for r in history}
    known_titles = {title_fingerprint(r.get("Project Title", "")) for r in history}
    known_urls.discard("")
    known_titles.discard("")

    fresh: list[dict] = []
    for row in new_rows:
        u = normalise_url(row.get("Source URL", ""))
        t = title_fingerprint(row.get("Project Title", ""))
        if (u and u in known_urls) or (t and t in known_titles):
            continue
        if u:
            known_urls.add(u)
        if t:
            known_titles.add(t)
        row = dict(row)
        row["First Seen"] = run_date
        fresh.append(row)

    merged = fresh + history

    # Rolling 12-month retention, based on when we first saw the row.
    cutoff = (today() - datetime.timedelta(days=HISTORY_RETENTION_DAYS)).isoformat()
    before = len(merged)
    merged = [r for r in merged if (r.get("First Seen") or run_date) >= cutoff]
    if before != len(merged):
        log.info("Retention: dropped %d rows first seen before %s",
                 before - len(merged), cutoff)

    log.info("History merge: %d new, %d carried forward, %d total",
             len(fresh), len(merged) - len(fresh), len(merged))
    return merged, fresh


def sort_for_report(rows: list[dict]) -> list[dict]:
    """
    Newest batch first, then by publish date descending, then by relevance.

    v1 sorted by source type first, which pushed all US procurement notices to
    the bottom of the sheet regardless of how recent or relevant they were.
    """
    def key(r: dict):
        return (
            _desc(r.get("First Seen") or ""),
            _desc(r.get("Date Published") or ""),
            -int(r.get("Score") or 0),
        )
    return sorted(rows, key=key)


def _desc(s: str) -> str:
    """Invert a date string so an ascending sort yields newest-first."""
    if not s:
        return " "          # empty sorts last
    return "".join(str(9 - int(c)) if c.isdigit() else c for c in s)


# ==============================================================================
# EXCEL REPORT
# ==============================================================================
#
# Column set is unchanged from v1 by design — Trevor's note was to improve the
# data in the existing columns, not to add new ones.

HEADERS = ["Project Title", "Source URL", "Summary", "Date Published", "Location"]
COL_WIDTHS = {"Project Title": 46, "Source URL": 44, "Summary": 68,
              "Date Published": 15, "Location": 28}

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL = PatternFill("solid", fgColor="DCE6F1")
NEW_ROW_FILL = PatternFill("solid", fgColor="FFF2CC")   # this week's finds
THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BODY_FONT = Font(name="Calibri", size=10)
LINK_FONT = Font(name="Calibri", size=10, color="0563C1", underline="single")


def write_sheet(ws, rows: list[dict], this_run: str) -> None:
    ws.freeze_panes = "A2"

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 20

    for idx, row in enumerate(rows, start=2):
        # Rows first seen in this run are highlighted so the team can tell at
        # a glance what is new in a cumulative sheet.
        is_new = (row.get("First Seen") or "") == this_run
        fill = NEW_ROW_FILL if is_new else (ALT_ROW_FILL if idx % 2 == 0 else PatternFill())
        for col, header in enumerate(HEADERS, start=1):
            value = row.get(header, "")
            cell = ws.cell(row=idx, column=col, value=value)
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(header in ("Project Title", "Summary")))
            cell.font = BODY_FONT
            if header == "Source URL" and value:
                # Excel rejects hyperlinks longer than 255 chars.
                if len(str(value)) <= 255:
                    cell.hyperlink = value
                cell.font = LINK_FONT

    for col, header in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = COL_WIDTHS[header]

    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows) + 1}"


def build_workbook(buckets: dict[str, list[dict]], this_run: str) -> openpyxl.Workbook:
    """Five tabs: USA, Canada, Australia-NZ, EMEA, Needs Review."""
    wb = openpyxl.Workbook()
    first = True
    for region in REGION_ORDER:
        rows = sort_for_report(buckets.get(region, []))
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = REGION_TAB_NAMES[region]
        write_sheet(ws, rows, this_run)
        ws.sheet_properties.tabColor = "C00000" if region == "REVIEW" else "1F3864"
    return wb


def workbook_to_bytes(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ==============================================================================
# EMAIL
# ==============================================================================
#
# Two transports. Graph is preferred: Microsoft is retiring Basic Auth for
# SMTP AUTH client submission in Exchange Online, so an app-password style
# SMTP setup on a safespill.com mailbox would work now and break later.
# If the GRAPH_* secrets are absent, this falls back to SMTP unchanged, so
# nothing breaks while IT is still setting up the app registration.

EMAIL_BODY = """Hi Safespill Team,
{notice_block}
Attached is this week's Safespill Hangar Intelligence Report.

NEW THIS WEEK: {new_count} items (highlighted in yellow in the spreadsheet)
TOTAL IN REPORT: {total_count} items, covering the last {retention_months} months

The report is cumulative — previous weeks' findings stay in the workbook, so
there is no need to dig back through older emails.

Tabs:
  USA .............. {usa} items
  Canada ........... {canada} items
  Australia-NZ ..... {anz} items
  EMEA ............. {emea} items
  Needs Review ..... {review} items (location could not be confirmed)

Sources: Google News, SAM.gov (US), CanadaBuys (CA), AusTender (AU),
NZ GETS (NZ), TED Europa (EU).

Safespill Automated Intelligence Report
"""


def build_quota_notice(quota: dict, error_count: int) -> str:
    if error_count > 0:
        return (
            f"WARNING: Coverage may be incomplete this week. The scraper hit "
            f"SerpAPI rate or quota limits {error_count} time(s), so some "
            f"searches returned nothing. Upgrading from the free SerpAPI plan "
            f"would resolve this."
        )
    if not quota:
        return ""
    try:
        left = int(quota.get("total_searches_left", 0))
        monthly = int(quota.get("searches_per_month", 0))
        if monthly <= 0:
            return ""
        pct_left = 100.0 * left / monthly
    except (ValueError, TypeError):
        return ""
    if pct_left < 10:
        return (f"WARNING: SerpAPI quota is critically low ({left} of {monthly} "
                f"searches left this month). Next week's report may be incomplete.")
    if pct_left < 25:
        return (f"Note: SerpAPI usage is at {100 - pct_left:.0f}% for the month "
                f"({left} searches left).")
    return ""


FALLBACK_RECIPIENT = "trevorw@safespill.com"


def recipients() -> list[str]:
    """
    Parse the REPORT_RECIPIENT secret into an address list.

    Tolerates commas, semicolons, newlines, padding and a trailing separator,
    because the value is pasted into a GitHub secret box by hand.

    If the secret is missing, empty, or contains nothing address-like, this
    falls back to a single known-good address and logs an ERROR. Returning an
    empty list here would make the run report success while silently emailing
    no one — the kind of failure nobody notices until someone asks where the
    report went.
    """
    raw = env("REPORT_RECIPIENT", "")
    parts = [a.strip() for a in re.split(r'[,;\s]+', raw) if a.strip()]
    valid = [a for a in parts if "@" in a and "." in a.split("@")[-1]]

    dropped = [p for p in parts if p not in valid]
    if dropped:
        log.warning("Ignoring malformed entries in REPORT_RECIPIENT: %s", dropped)

    if not valid:
        log.error(
            "REPORT_RECIPIENT is empty or unparseable (got %r). Falling back to "
            "%s. Check the REPORT_RECIPIENT repository secret exists and is "
            "spelled exactly that way.", raw, FALLBACK_RECIPIENT)
        return [FALLBACK_RECIPIENT]

    # De-duplicate case-insensitively while preserving the given order.
    seen, unique = set(), []
    for a in valid:
        if a.lower() not in seen:
            seen.add(a.lower())
            unique.append(a)
    log.info("Report recipients (%d): %s", len(unique), ", ".join(unique))
    return unique


def graph_configured() -> bool:
    return all(env(k) for k in
               ("GRAPH_TENANT_ID", "GRAPH_CLIENT_ID", "GRAPH_CLIENT_SECRET", "GRAPH_SENDER"))


def graph_access_token() -> str:
    url = f"https://login.microsoftonline.com/{env('GRAPH_TENANT_ID')}/oauth2/v2.0/token"
    data = {
        "client_id": env("GRAPH_CLIENT_ID"),
        "client_secret": env("GRAPH_CLIENT_SECRET"),
        "scope": "https://graph.microsoft.com/.default",
        "grant_type": "client_credentials",
    }
    r = requests.post(url, data=data, timeout=30)
    r.raise_for_status()
    return r.json()["access_token"]


def send_email_graph(xlsx: bytes, filename: str, subject: str, body: str) -> None:
    token = graph_access_token()
    sender = env("GRAPH_SENDER")
    payload = {
        "message": {
            "subject": subject,
            "body": {"contentType": "Text", "content": body},
            "toRecipients": [{"emailAddress": {"address": a}} for a in recipients()],
            "attachments": [{
                "@odata.type": "#microsoft.graph.fileAttachment",
                "name": filename,
                "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "contentBytes": base64.b64encode(xlsx).decode("ascii"),
            }],
        },
        "saveToSentItems": True,
    }
    r = requests.post(
        f"https://graph.microsoft.com/v1.0/users/{sender}/sendMail",
        headers={"Authorization": "Bearer " + token,
                 "Content-Type": "application/json"},
        json=payload, timeout=120)
    if r.status_code not in (200, 202):
        raise RuntimeError(f"Graph sendMail failed: {r.status_code} {r.text[:400]}")
    log.info("Email sent via Microsoft Graph from %s to %s", sender, recipients())


def send_email_smtp(xlsx: bytes, filename: str, subject: str, body: str) -> None:
    import smtplib
    from email import encoders
    from email.mime.base import MIMEBase
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    user = env("SMTP_USER")
    msg = MIMEMultipart()
    msg["From"] = env("SMTP_FROM", user)
    msg["To"] = ", ".join(recipients())
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    part = MIMEBase("application",
                    "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.set_payload(xlsx)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)

    with smtplib.SMTP(env("SMTP_HOST", "smtp.gmail.com"),
                      int(env("SMTP_PORT", "587"))) as server:
        server.ehlo()
        server.starttls()
        server.login(user, env("SMTP_PASSWORD"))
        server.sendmail(user, recipients(), msg.as_string())
    log.info("Email sent via SMTP to %s", recipients())


def send_report(xlsx: bytes, filename: str, subject: str, body: str) -> None:
    if graph_configured():
        try:
            send_email_graph(xlsx, filename, subject, body)
            return
        except Exception as exc:
            log.error("Graph send failed (%s) — falling back to SMTP.", exc)
    if env("SMTP_USER") and env("SMTP_PASSWORD"):
        send_email_smtp(xlsx, filename, subject, body)
    else:
        log.error("No working email transport configured. Report NOT sent.")


# ==============================================================================
# MAIN
# ==============================================================================

def collect_rows() -> list[dict]:
    """Run every source and return the raw pooled rows."""
    news_start, news_end = date_range(LOOKBACK_DAYS)
    proc_start, proc_end = date_range(PROCUREMENT_LOOKBACK_DAYS)
    all_rows: list[dict] = []

    # 1. Google News
    week_iso = today().isocalendar()[1]
    queries = queries_for_this_week(week_iso)
    log.info("Running %d SerpAPI news searches (week %d rotation)", len(queries), week_iso)
    for q in queries:
        for item in serpapi_news_search(q):
            row = parse_news_result(item)
            if row:
                all_rows.append(row)
        time.sleep(1.2)          # stay under the 50/hour free-tier cap

    # 2. SAM.gov (United States)
    sam_start, sam_end = date_range(SAM_LOOKBACK_DAYS)
    log.info("[USA] SAM.gov, past %d days", SAM_LOOKBACK_DAYS)
    all_rows.extend(samgov_search(sam_start, sam_end))

    # 3. CanadaBuys
    log.info("[CANADA] CanadaBuys open data")
    all_rows.extend(canadabuys_search(proc_start))

    # 4. AusTender
    log.info("[ANZ] AusTender OCDS")
    all_rows.extend(austender_search(proc_start, proc_end))

    # 5. NZ GETS
    log.info("[ANZ] New Zealand GETS RSS")
    all_rows.extend(nz_gets_search())

    # 6. TED Europa
    log.info("[EMEA] TED Europa")
    all_rows.extend(ted_europa_search(proc_start, proc_end))

    return all_rows


def filter_rows(rows: list[dict]) -> list[dict]:
    """Apply the blocklist, vendor-page test, and relevance threshold."""
    kept, stats = [], {"blocked": 0, "vendor": 0, "low_score": 0}
    for row in rows:
        url = row.get("Source URL", "")
        title = row.get("Project Title", "")
        if row.get("Source") == "news":
            if is_blocked_domain(url):
                stats["blocked"] += 1
                continue
            if looks_like_vendor_page(url, title):
                stats["vendor"] += 1
                continue
        score = relevance_score(row)
        if score < MIN_RELEVANCE_SCORE:
            stats["low_score"] += 1
            continue
        row["Score"] = score
        kept.append(row)
    log.info("Relevance filter: kept %d | dropped %d blocked-domain, "
             "%d vendor-page, %d below score %d",
             len(kept), stats["blocked"], stats["vendor"],
             stats["low_score"], MIN_RELEVANCE_SCORE)
    return kept


def filter_dates(rows: list[dict]) -> list[dict]:
    """
    Drop rows with no credible publish date, or one outside the window.
    v1 had no date filter at all: 171 of 341 rows (50%) in the 8/3 report
    predated its own 7-day lookback, some by nearly a year.
    """
    kept, dropped = [], 0
    for row in rows:
        src = row.get("Source", "")
        max_age = SAM_LOOKBACK_DAYS if src != "news" else LOOKBACK_DAYS
        if date_is_sane(row.get("Date Published", ""), max_age):
            kept.append(row)
        else:
            dropped += 1
    log.info("Date filter: kept %d, dropped %d (missing/stale/future dates)",
             len(kept), dropped)
    return kept


def trim_to_target(rows: list[dict]) -> list[dict]:
    """
    If a week over-delivers, keep the highest-scoring rows so the sheet stays
    at a readable size. Under-delivering is left alone — a quiet week is real
    information, and padding it with junk defeats the purpose.
    """
    if len(rows) <= TARGET_NEW_ROWS_MAX:
        return rows
    ranked = sorted(rows, key=lambda r: -int(r.get("Score") or 0))
    log.info("Trimming %d new rows to top %d by relevance score",
             len(rows), TARGET_NEW_ROWS_MAX)
    return ranked[:TARGET_NEW_ROWS_MAX]


def main() -> None:
    global _SERPAPI_ERRORS, _SERPAPI_CALLS
    _SERPAPI_ERRORS = 0
    _SERPAPI_CALLS = 0

    run_date = today().isoformat()
    week = week_label()
    log.info("=== Safespill Hangar Report run %s (week of %s) ===", run_date, week)

    starting_quota = check_serpapi_quota()
    log_quota_status("Start", starting_quota)

    # 1. Collect
    raw = collect_rows()
    log.info("Collected %d raw rows", len(raw))

    # 2. Dedupe within this run
    rows = deduplicate(raw)
    log.info("After in-run dedupe: %d", len(rows))

    # 3. Relevance filter BEFORE page enrichment — enrichment is the slow
    #    part, so there is no sense fetching pages we are going to discard.
    rows = filter_rows(rows)

    # 4. Drop rows already in history, so we do not spend fetches on them
    history = load_history()
    known_urls = {normalise_url(r.get("Source URL", "")) for r in history}
    known_titles = {title_fingerprint(r.get("Project Title", "")) for r in history}
    known_urls.discard("")
    known_titles.discard("")
    before = len(rows)
    rows = [r for r in rows
            if normalise_url(r.get("Source URL", "")) not in known_urls
            and title_fingerprint(r.get("Project Title", "")) not in known_titles]
    log.info("Dropped %d rows already in history", before - len(rows))

    # 5. Enrich with real dates and locations from the article pages
    rows = enrich_with_page_data(rows)

    # 6. Date sanity filter
    rows = filter_dates(rows)

    # 7. Trim to the weekly target
    rows = trim_to_target(rows)

    # 8. Merge into cumulative history
    merged, fresh = merge_history(history, rows, run_date)
    save_history(merged)

    # 9. Split into tabs and build the workbook
    buckets = split_by_region(merged)
    for region in REGION_ORDER:
        log.info("  %-14s %d rows", REGION_TAB_NAMES[region], len(buckets[region]))

    wb = build_workbook(buckets, run_date)
    xlsx = workbook_to_bytes(wb)
    filename = f"Safespill_Hangar_Report_{week}.xlsx"

    # 10. Email
    ending_quota = check_serpapi_quota()
    log_quota_status("End", ending_quota)

    body = EMAIL_BODY.format(
        notice_block=("\n" + build_quota_notice(ending_quota, _SERPAPI_ERRORS) + "\n"
                      if build_quota_notice(ending_quota, _SERPAPI_ERRORS) else ""),
        new_count=len(fresh),
        total_count=len(merged),
        retention_months=HISTORY_RETENTION_DAYS // 30,
        usa=len(buckets["USA"]), canada=len(buckets["CANADA"]),
        anz=len(buckets["ANZ"]), emea=len(buckets["EMEA"]),
        review=len(buckets["REVIEW"]),
    )
    subject = f"Safespill Hangar Intelligence Report - Week of {week}"
    send_report(xlsx, filename, subject, body)

    log.info("SerpAPI searches used this run: %d | errors: %d",
             _SERPAPI_CALLS, _SERPAPI_ERRORS)
    log.info("Done. %d new, %d total in report.", len(fresh), len(merged))


if __name__ == "__main__":
    main()
